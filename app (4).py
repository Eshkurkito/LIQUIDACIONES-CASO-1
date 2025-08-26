
import streamlit as st
import pandas as pd
import numpy as np
from datetime import date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

st.set_page_config(page_title="LIQUIDACIONES (Casos 1-5)", page_icon="📊", layout="wide")

# =====================
# Utilidades
# =====================
def ensure_required(df, required, ctx=""):
    missing = [c for c in required if c not in df.columns]
    if missing:
        st.error(f"Faltan columnas requeridas: {missing} {('en ' + ctx) if ctx else ''}. "
                 "Comprueba el Excel (la cabecera puede estar en la fila 1).")
        st.stop()

def _first_existing(df, candidates):
    norm_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        k = str(cand).strip().lower()
        if k in norm_map:
            return norm_map[k]
    return None

def normalize_columns(df):
    out = df.copy()
    col_aloj = _first_existing(out, ["Nombre alojamiento","Alojamiento","Nombre del alojamiento","Nombre Alojamiento"])
    col_fent = _first_existing(out, ["Fecha entrada","Fecha de entrada"])
    col_fsal = _first_existing(out, ["Fecha salida","Fecha de salida"])
    col_noch = _first_existing(out, ["noches","Noches","Noches ocupadas"])
    col_alq  = _first_existing(out, ["Alquiler con tasas","Ingreso alojamiento","Importe alojamiento"])
    col_ext  = _first_existing(out, ["Extras con tasas","Ingreso limpieza","Limpieza","Importe limpieza"])
    col_tot  = _first_existing(out, ["Total reserva con tasas","Total ingresos","Total"])
    col_port = _first_existing(out, ["Web origen","Portal","Canal","Fuente"])
    col_comi = _first_existing(out, ["Comisión Portal/Intermediario: Comisión calculada","Comisión portal","Comisión"])
    col_ivaal= _first_existing(out, ["IVA del alojamiento","IVA alojamiento"])

    rename = {}
    if col_aloj: rename[col_aloj] = "Alojamiento"
    if col_fent: rename[col_fent] = "Fecha entrada"
    if col_fsal: rename[col_fsal] = "Fecha salida"
    if col_noch: rename[col_noch] = "Noches ocupadas"
    if col_alq:  rename[col_alq]  = "Ingreso alojamiento"
    if col_ext:  rename[col_ext]  = "Ingreso limpieza"
    if col_tot:  rename[col_tot]  = "Total ingresos"
    if col_port: rename[col_port] = "Portal"
    if col_comi: rename[col_comi] = "Comisión portal"
    if col_ivaal:rename[col_ivaal]= "IVA del alquiler"

    out.rename(columns=rename, inplace=True)

    if "Alojamiento" not in out.columns:
        candidates = [c for c in out.columns if "aloj" in str(c).lower()]
        if candidates:
            out.rename(columns={candidates[0]: "Alojamiento"}, inplace=True)

    for c in ["Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisión portal","IVA del alquiler","Noches ocupadas"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)

    for c in ["Fecha entrada","Fecha salida"]:
        if c in out.columns:
            out[c] = pd.to_datetime(out[c], errors="coerce", dayfirst=True)

    if "Alojamiento" in out.columns:
        out["Alojamiento"] = out["Alojamiento"].astype(str).str.strip().str.upper()

    return out

def keep_dates_text(df, cols):
    out = df.copy()
    for c in cols:
        if c in out.columns:
            out[c] = pd.to_datetime(out[c], errors="coerce", dayfirst=True).dt.strftime("%d/%m/%Y").fillna("")
    return out

def apply_booking_vat_on_commission(df, portal_col="Portal", commission_col="Comisión portal"):
    out = df.copy()
    if portal_col in out.columns and commission_col in out.columns:
        ser = out[portal_col]
        # Forzar a serie de texto segura
        try:
            ser = ser.astype("string")
        except Exception:
            ser = ser.astype(str)
        mask = ser.fillna("").astype(str).str.lower().str.contains("booking", na=False)
        # Asegurar que la comisión es numérica
        out[commission_col] = pd.to_numeric(out[commission_col], errors="coerce").fillna(0.0)
        out.loc[mask, commission_col] = out.loc[mask, commission_col] * 1.21
    return out

def to_excel_grouped(df, group_col="Alojamiento", name="Liquidación"):
    number_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]

    # Columna vacía antes de "Pago recibido"
    if "Pago recibido" in df.columns and "" not in df.columns:
        idx = list(df.columns).index("Pago recibido")
        df = df.copy()
        df.insert(idx, "", "")

    wb = Workbook()
    ws = wb.active
    ws.title = name

    bold = Font(bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center")

    cols = df.columns.tolist()
    row_cursor = 1

    for g, sub in df.groupby(group_col):
        # Encabezado
        for j, cn in enumerate(cols, start=1):
            c = ws.cell(row=row_cursor, column=j, value=cn)
            c.font = bold; c.border = border; c.alignment = center
        row_cursor += 1

        # Filas
        for _, r in sub.iterrows():
            for j, cn in enumerate(cols, start=1):
                c = ws.cell(row=row_cursor, column=j, value=r[cn])
                c.border = border
            row_cursor += 1

        # Subtotales
        for j, cn in enumerate(cols, start=1):
            if cn == group_col:
                ws.cell(row=row_cursor, column=j, value="Total").font = bold
            elif cn in number_cols:
                val = float(round(sub[cn].sum(),2))
                c = ws.cell(row=row_cursor, column=j, value=val)
                c.font = bold; c.border = border
        row_cursor += 2

    # Formato numérico español
    nf = "#.##0,00"
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in row:
            if isinstance(c.value, (int,float)):
                c.number_format = nf

    # Auto ancho
    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max(10, min(max_len+2, 60))

    buf = BytesIO(); wb.save(buf); buf.seek(0); return buf

# =====================
# Configuración de casos (reglas memorizadas)
# =====================
case1_percent_amenities = {
    "APOLO 180": (0.20, 12.04), "ALMIRANTE 01": (0.22, 11.33), "ALMIRANTE 02": (0.22, 11.33),
    "CADIZ": (0.20, 9.11), "DENIA 61": (0.20, 10.96), "DOLORES ALCAYDE 04": (0.20, 11.33),
    "DR.LLUCH": (0.20, 11.16), "ERUDITO": (0.20, 13.37), "GOZALBO": (0.20, 15.25),
    "LA ELIANA": (0.20, 15.25), "MORAIRA": (0.25, 11.33), "NAPOLES Y SICILIA": (0.25, 0.00),
    "OLIVERETA 5": (0.20, 0.00), "OVE 01": (0.18, 0.00), "OVE 02": (0.18, 0.00),
    "QUART I": (0.20, 9.09), "QUART II": (0.20, 9.09), "SAN LUIS": (0.20, 11.02),
    "SERRANOS": (0.20, 13.37), "SEVILLA": (0.18, 9.45), "TUNDIDORES": (0.20, 7.85),
    "VALLE": (0.20, 11.33),
}
case1_props = set(case1_percent_amenities.keys())

case2_percent_amenities = {
    "VISITACION": (0.20, 14.88),
    "PADRE PORTA 6": (0.20, 12.09), "PADRE PORTA 7": (0.20, 12.09), "PADRE PORTA 8": (0.20, 12.09),
    "PADRE PORTA 9": (0.20, 12.09), "PADRE PORTA 10": (0.20, 12.09),
    "LLADRO Y MALLI 00": (0.20, 9.45), "LLADRO Y MALLI 01": (0.20, 9.45), "LLADRO Y MALLI 02": (0.20, 9.45),
    "LLADRO Y MALLI 03": (0.20, 9.45), "LLADRO Y MALLI 04": (0.20, 9.45),
    "APOLO 29": (0.20, 11.58), "APOLO 197": (0.20, 17.40),
}
case2_props = set(case2_percent_amenities.keys())

case3_cleaning_amenities = {
    "ZAPATEROS 10-2": (0.20, 60.00, 15.24),
    "ZAPATEROS 10-6": (0.20, 75.00, 15.24),
    "ZAPATEROS 10-8": (0.20, 75.00, 15.24),
    "ZAPATEROS 12-5": (0.20, 60.00, 11.33),
    "ALFARO": (0.20, 80.00, 14.88),
}
case3_props = set(case3_cleaning_amenities.keys())

case4_props = {
    "SERRERIA 04", "SERRERIA 05", "RETOR A", "RETOR B",
    "PASAJE ANGELES Y FEDERICO 01", "PASAJE ANGELES Y FEDERICO 02", "PASAJE ANGELES Y FEDERICO 03",
    "MALILLA 05", "MALILLA 06", "MALILLA 07", "MALILLA 08", "MALILLA 14", "MALILLA 15",
    "BENICALAP 01", "BENICALAP 02", "BENICALAP 03", "BENICALAP 04", "BENICALAP 05", "BENICALAP 06"
}
case5_percent_amenities = {
    "HOMERO 01": (0.20, 0.00), "HOMERO 02": (0.20, 0.00),
    "CARCAIXENT 01": (0.20, 8.60), "CARCAIXENT 02": (0.20, 8.60),
}
case5_props = set(case5_percent_amenities.keys())

# =====================
# Procesadores (cálculo por caso)
# =====================
def process_case1(df):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisión portal","Portal"], "process_case1")
    df = apply_booking_vat_on_commission(df, "Portal", "Comisión portal")

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = case1_percent_amenities.get(key,(0.20,0))[0]
        return float(r.get("Ingreso alojamiento",0))*pct*1.21

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case1_percent_amenities.get(key,(0.20,0))[1])

    out = df.copy()
    out["Honorarios Florit"] = out.apply(honorarios,axis=1)
    out["Gasto limpieza"] = out.get("Ingreso limpieza",0.0)
    out["Amenities"] = out.apply(amenities,axis=1)
    out["Total Gastos"] = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"] = out["Total ingresos"] - out["Comisión portal"]
    out = keep_dates_text(out, ["Fecha entrada","Fecha salida"])
    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","Ingreso limpieza","Total ingresos","Portal","Comisión portal","Honorarios Florit","Gasto limpieza","Amenities","Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].sort_values(by=["Alojamiento","Fecha entrada"] if "Fecha entrada" in cols else ["Alojamiento"])

def process_case2(df):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Total ingresos","Comisión portal","Portal"], "process_case2")
    out = df.copy()
    if {"Alojamiento","Portal","Comisión portal"}.issubset(out.columns):
        mask_apolo = out["Alojamiento"].astype(str).str.upper().isin({"APOLO 29","APOLO 197"})
        mask_book = out["Portal"].astype(str).str.lower().str.contains("booking", na=False)
        out.loc[mask_apolo & mask_book,"Comisión portal"] = out.loc[mask_apolo & mask_book,"Comisión portal"]*1.21

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = case2_percent_amenities.get(key,(0.20,0))[0]
        base = float(r.get("Ingreso alojamiento",0)) - float(r.get("IVA del alquiler",0))
        return base*pct*1.21

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case2_percent_amenities.get(key,(0.20,0))[1])

    out["Honorarios Florit"] = out.apply(honorarios,axis=1)
    out["Gasto limpieza"] = out.get("Ingreso limpieza",0.0)
    out["Amenities"] = out.apply(amenities,axis=1)
    out["Total Gastos"] = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"] = out["Total ingresos"] - out["Comisión portal"]
    out = keep_dates_text(out, ["Fecha entrada","Fecha salida"])
    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","IVA del alquiler","Ingreso limpieza","Total ingresos","Portal","Comisión portal","Honorarios Florit","Gasto limpieza","Amenities","Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].sort_values(by=["Alojamiento","Fecha entrada"] if "Fecha entrada" in cols else ["Alojamiento"])

def process_case3(df):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Total ingresos","Comisión portal","Portal"], "process_case3")
    out = apply_booking_vat_on_commission(df, "Portal", "Comisión portal")

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = case3_cleaning_amenities.get(key,(0.20,None,None))[0]
        base = float(r.get("Ingreso alojamiento",0)) - float(r.get("Comisión portal",0))
        return base*pct*1.21

    def gasto_limpieza(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case3_cleaning_amenities.get(key,(0.20,0,0))[1])

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case3_cleaning_amenities.get(key,(0.20,0,0))[2])

    out = out.copy()
    out["Honorarios Florit"] = out.apply(honorarios,axis=1)
    out["Gasto limpieza"] = out.apply(gasto_limpieza,axis=1)
    out["Amenities"] = out.apply(amenities,axis=1)
    out["Total Gastos"] = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"] = out["Total ingresos"] - out["Comisión portal"]
    out = keep_dates_text(out, ["Fecha entrada","Fecha salida"])
    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","Ingreso limpieza","Total ingresos","Portal","Comisión portal","Honorarios Florit","Gasto limpieza","Amenities","Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].sort_values(by=["Alojamiento","Fecha entrada"] if "Fecha entrada" in cols else ["Alojamiento"])

def process_case4(df):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Comisión portal"], "process_case4")
    out = df.copy()

    def honorarios(r):
        base = float(r.get("Ingreso alojamiento",0)) - float(r.get("IVA del alquiler",0)) - float(r.get("Comisión portal",0))
        return base*0.20

    out["Honorarios Florit"] = out.apply(honorarios,axis=1)
    out["Gasto limpieza"] = out.get("Ingreso limpieza",0.0)
    out["Amenities"] = 0.0
    out["Total Gastos"] = out[["Comisión portal","Honorarios Florit"]].sum(axis=1) + out["Gasto limpieza"] + out["Amenities"]
    out["Pago al propietario"] = out["Ingreso alojamiento"] - out.get("IVA del alquiler",0.0) - out["Comisión portal"] - out["Honorarios Florit"]
    out["Pago recibido"] = out["Ingreso alojamiento"] + out.get("Ingreso limpieza",0.0) - out["Comisión portal"]
    out = keep_dates_text(out, ["Fecha entrada","Fecha salida"])
    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","IVA del alquiler","Portal","Comisión portal","Honorarios Florit","Pago al propietario","Pago recibido","Ingreso limpieza"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].sort_values(by=["Alojamiento","Fecha entrada"] if "Fecha entrada" in cols else ["Alojamiento"])

def process_case5(df):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Total ingresos","Comisión portal"], "process_case5")
    out = df.copy()

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = case5_percent_amenities.get(key,(0.20,0))[0]
        base = float(r.get("Ingreso alojamiento",0)) - float(r.get("IVA del alquiler",0)) - float(r.get("Comisión portal",0))
        return base*pct*1.21

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case5_percent_amenities.get(key,(0.20,0))[1])

    out["Honorarios Florit"] = out.apply(honorarios,axis=1)
    out["Gasto limpieza"] = out.get("Ingreso limpieza",0.0)
    out["Amenities"] = out.apply(amenities,axis=1)
    out["Total Gastos"] = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"] = out["Total ingresos"] - out["Comisión portal"]
    out = keep_dates_text(out, ["Fecha entrada","Fecha salida"])
    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","IVA del alquiler","Ingreso limpieza","Total ingresos","Portal","Comisión portal","Honorarios Florit","Gasto limpieza","Amenities","Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].sort_values(by=["Alojamiento","Fecha entrada"] if "Fecha entrada" in cols else ["Alojamiento"])

processors = {1: process_case1, 2: process_case2, 3: process_case3, 4: process_case4, 5: process_case5}

def detect_case(df):
    df_norm = normalize_columns(df)
    alos = set(df_norm.get("Alojamiento", pd.Series(dtype=str)).astype(str).str.upper().unique())
    scores = {
        1: len(alos & set(case1_props)),
        2: len(alos & set(case2_props)),
        3: len(alos & set(case3_props)),
        4: len(alos & set(case4_props)),
        5: len(alos & set(case5_props)),
    }
    best = max(scores, key=scores.get)
    return best, scores

# =====================
# Período y prorrateo
# =====================
def overlap_nights(ci, co, start, end):
    if pd.isna(ci) or pd.isna(co): return 0
    start = pd.to_datetime(start); end = pd.to_datetime(end) + pd.Timedelta(days=1)
    a = max(ci, start); b = min(co, end)
    return max(0, (b-a).days)

def apply_period(df, start_date, end_date, prorate=True, limpieza_mode="prorratear", amenities_mode="prorratear"):
    out = df.copy()
    fe = pd.to_datetime(out.get("Fecha entrada"), errors="coerce", dayfirst=True)
    fs = pd.to_datetime(out.get("Fecha salida"), errors="coerce", dayfirst=True)
    tot_noches = pd.to_numeric(out.get("Noches ocupadas"), errors="coerce").fillna(0)
    nights = [overlap_nights(a,b,start_date,end_date) for a,b in zip(fe,fs)]
    out["Noches periodo"] = pd.Series(nights, index=out.index).astype(float)

    if not prorate:
        mask = out["Noches periodo"] > 0
        out = out[mask].drop(columns=["Noches periodo"])
        out["Fecha entrada"] = fe.dt.strftime("%d/%m/%Y")
        out["Fecha salida"] = fs.dt.strftime("%d/%m/%Y")
        return out

    ratio = np.where(tot_noches.to_numpy()>0, (out["Noches periodo"]/tot_noches).to_numpy(), 0.0)

    for col in ["Ingreso alojamiento","Total ingresos","Comisión portal","Honorarios Florit"]:
        if col in out.columns:
            out[col] = (pd.to_numeric(out[col], errors="coerce").fillna(0).to_numpy()*ratio).round(2)

    if {"Ingreso limpieza","Gasto limpieza"}.issubset(out.columns):
        if limpieza_mode=="prorratear":
            out["Ingreso limpieza"] = (pd.to_numeric(out["Ingreso limpieza"], errors="coerce").fillna(0).to_numpy()*ratio).round(2)
            out["Gasto limpieza"] = out["Ingreso limpieza"]
        elif limpieza_mode=="salida":
            mask = (fs >= pd.to_datetime(start_date)) & (fs <= pd.to_datetime(end_date))
            out["Ingreso limpieza"] = np.where(mask, out["Ingreso limpieza"], 0.0)
            out["Gasto limpieza"] = out["Ingreso limpieza"]
        elif limpieza_mode=="entrada":
            mask = (fe >= pd.to_datetime(start_date)) & (fe <= pd.to_datetime(end_date))
            out["Ingreso limpieza"] = np.where(mask, out["Ingreso limpieza"], 0.0)
            out["Gasto limpieza"] = out["Ingreso limpieza"]

    if "Amenities" in out.columns:
        if amenities_mode=="prorratear":
            out["Amenities"] = (pd.to_numeric(out["Amenities"], errors="coerce").fillna(0).to_numpy()*ratio).round(2)
        elif amenities_mode=="salida":
            mask = (fs >= pd.to_datetime(start_date)) & (fs <= pd.to_datetime(end_date))
            out["Amenities"] = np.where(mask, out["Amenities"], 0.0)
        elif amenities_mode=="entrada":
            mask = (fe >= pd.to_datetime(start_date)) & (fe <= pd.to_datetime(end_date))
            out["Amenities"] = np.where(mask, out["Amenities"], 0.0)

    for c in ["Total Gastos","Pago al propietario","Pago recibido"]:
        if c in out.columns: out.drop(columns=[c], inplace=True, errors="ignore")
    if {"Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"}.issubset(out.columns):
        out["Total Gastos"] = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1).round(2)
    if {"Total ingresos","Total Gastos"}.issubset(out.columns):
        out["Pago al propietario"] = (out["Total ingresos"] - out["Total Gastos"]).round(2)
    if {"Total ingresos","Comisión portal"}.issubset(out.columns):
        out["Pago recibido"] = (out["Total ingresos"] - out["Comisión portal"]).round(2)

    out["Fecha entrada"] = fe.dt.strftime("%d/%m/%Y")
    out["Fecha salida"] = fs.dt.strftime("%d/%m/%Y")
    return out.drop(columns=["Noches periodo"], errors="ignore")

# =====================
# Detección de cabecera y UI
# =====================
def autodetect_header_row(df_noheader, max_rows_to_check=8):
    expected = {"nombre","alojamiento","fecha entrada","fecha salida","noches","alquiler con tasas","extras con tasas","total reserva con tasas","comisión","portal","web origen","iva"}
    best, best_score = 0, -1
    for i in range(min(max_rows_to_check, len(df_noheader))):
        row = df_noheader.iloc[i].astype(str).str.lower().str.strip()
        score = sum(any(token in cell for cell in row) for token in expected)
        if score > best_score:
            best, best_score = i, score
    return best

st.title("📊 LIQUIDACIONES Automáticas (Casos 1–5)")
st.caption("Sube Excel de Avantio (.xlsx) • Detecta caso • Período y prorrateo • Exporta Excel formateado")

with st.sidebar:
    st.header("Período a liquidar")
    col1, col2 = st.columns(2)
    with col1: start_date = st.date_input("Desde", value=date(date.today().year, date.today().month, 1))
    with col2: end_date   = st.date_input("Hasta",  value=date(date.today().year, date.today().month, 28))
    st.divider()
    prorate = st.checkbox("Prorratear por noches en período", value=True)
    limpieza_mode = st.selectbox("Limpieza", ["prorratear","salida","entrada"], index=0)
    amenities_mode = st.selectbox("Amenities", ["prorratear","salida","entrada"], index=0)

file = st.file_uploader("Sube el archivo de reservas (.xlsx)", type=["xlsx"])

if file is not None:
    df_noheader = pd.read_excel(file, header=None)
    header_guess = autodetect_header_row(df_noheader)
    header_row = st.number_input("Fila de cabecera (1 = primera fila en Excel)", 1, 10, int(header_guess+1), 1) - 1

    df_in = pd.read_excel(file, header=header_row)
    st.write("Vista previa (primeras 12 filas):")
    st.dataframe(df_in.head(12), use_container_width=True)

    detected_case, scores = detect_case(df_in)
    st.info(f"**Caso detectado automáticamente: {detected_case}** | Recuento por caso: {scores}")
    case_choice = st.selectbox("Elegir caso (manual opcional)", [1,2,3,4,5], index=[1,2,3,4,5].index(detected_case))

    if st.button("Generar liquidación"):
        processor = processors[case_choice]
        base_df = processor(df_in)
        final_df = apply_period(base_df, start_date, end_date, prorate=prorate, limpieza_mode=limpieza_mode, amenities_mode=amenities_mode)

        st.success(f"Liquidación generada (Caso {case_choice}) • {start_date.strftime('%d/%m/%Y')}–{end_date.strftime('%d/%m/%Y')}")
        st.dataframe(final_df.head(50), use_container_width=True)

        excel_buffer = to_excel_grouped(final_df, group_col="Alojamiento", name=f"CASO {case_choice}")
        st.download_button("📥 Descargar Excel formateado", data=excel_buffer,
            file_name=f"Liquidacion_CASO_{case_choice}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
else:
    st.info("Sube un archivo para comenzar.")
