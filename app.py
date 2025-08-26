
import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="LIQUIDACIONES (Casos 1-5)", page_icon="📊", layout="wide")

# ----------------------
# Utilidades
# ----------------------
def to_excel_grouped(df, group_col="Alojamiento", name="Liquidación", number_cols=None):
    """
    Exporta un DataFrame a Excel agrupando por 'group_col', con encabezado repetido por grupo,
    bordes finos, encabezados en negrita, subtotales al final del grupo para todas las columnas numéricas
    y formato español (#.##0,00). Devuelve BytesIO listo para descarga.
    """
    if number_cols is None:
        number_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]

    # Insertar columna vacía antes de "Pago recibido" si no existe
    if "Pago recibido" in df.columns and "" not in df.columns:
        idx = list(df.columns).index("Pago recibido")
        df = df.copy()
        df.insert(idx, "", "")

    wb = Workbook()
    ws = wb.active
    ws.title = name

    bold = Font(bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center")

    cols = df.columns.tolist()
    row_cursor = 1

    for group_val, subdf in df.groupby(group_col):
        # Encabezado
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=row_cursor, column=col_idx, value=col_name)
            cell.font = bold
            cell.border = border
            cell.alignment = center
        row_cursor += 1

        # Filas
        for _, row in subdf.iterrows():
            for col_idx, col_name in enumerate(cols, start=1):
                value = row[col_name]
                cell = ws.cell(row=row_cursor, column=col_idx, value=value)
                cell.border = border
            row_cursor += 1

        # Subtotales para columnas numéricas
        for col_idx, col_name in enumerate(cols, start=1):
            if col_name == group_col:
                ws.cell(row=row_cursor, column=col_idx, value="Total").font = bold
            elif col_name in number_cols:
                subtotal = subdf[col_name].sum()
                c = ws.cell(row=row_cursor, column=col_idx, value=float(round(subtotal, 2)))
                c.font = bold
                c.border = border
        row_cursor += 2

    # Formato numérico español
    number_format = "#.##0,00"
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value is not None:
                cell.number_format = number_format

    # Auto ancho
    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max(10, min(max_len + 2, 60))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def keep_original_dates(df, date_cols):
    """Preserva las fechas como texto exactamente como vienen en origen."""
    out = df.copy()
    for c in date_cols:
        if c in out.columns:
            out[c] = out[c].astype(str)
    return out

# ----------------------
# Configuración por casos (según tus especificaciones definitivas)
# ----------------------
case1_percent_amenities = {
    "APOLO 180": (0.20, 12.04), "ALMIRANTE 01": (0.22, 11.33), "ALMIRANTE 02": (0.22, 11.33),
    "CADIZ": (0.20, 9.11), "DENIA 61": (0.20, 10.96), "DOLORES ALCAYDE 04": (0.20, 11.33),
    "DR.LLUCH": (0.20, 11.16), "ERUDITO": (0.20, 13.37), "GOZALBO": (0.20, 15.25),
    "LA ELIANA": (0.20, 15.25), "MORAIRA": (0.25, 11.33), "NAPOLES Y SICILIA": (0.25, 0.00),
    "OLIVERETA 5": (0.20, 0.00), "OVE 01": (0.18, 0.00), "OVE 02": (0.18, 0.00),
    "QUART I": (0.20, 9.09), "QUART II": (0.20, 9.09), "SAN LUIS": (0.20, 11.02),
    "SERRANOS": (0.20, 13.37), "SEVILLA": (0.18, 9.45), "TUNDIDORES": (0.20, 7.85),
    "VALLE": (0.20, 11.33),
}
case1_props = set(case1_percent_amenities.keys())

case2_percent_amenities = {
    "VISITACION": (0.20, 14.88),
    "PADRE PORTA 6": (0.20, 12.09), "PADRE PORTA 7": (0.20, 12.09), "PADRE PORTA 8": (0.20, 12.09),
    "PADRE PORTA 9": (0.20, 12.09), "PADRE PORTA 10": (0.20, 12.09),
    "LLADRO Y MALLI 00": (0.20, 9.45), "LLADRO Y MALLI 01": (0.20, 9.45), "LLADRO Y MALLI 02": (0.20, 9.45),
    "LLADRO Y MALLI 03": (0.20, 9.45), "LLADRO Y MALLI 04": (0.20, 9.45),
    "APOLO 29": (0.20, 11.58), "APOLO 197": (0.20, 17.40),
}
case2_props = set(case2_percent_amenities.keys())

case3_cleaning_amenities = {
    "ZAPATEROS 10-2": (0.20, 60.00, 15.24),
    "ZAPATEROS 10-6": (0.20, 75.00, 15.24),
    "ZAPATEROS 10-8": (0.20, 75.00, 15.24),
    "ZAPATEROS 12-5": (0.20, 60.00, 11.33),
    "ALFARO": (0.20, 80.00, 14.88),
}
case3_props = set(case3_cleaning_amenities.keys())

case4_props = {
    "SERRERIA 04", "SERRERIA 05", "RETOR A", "RETOR B",
    "PASAJE ANGELES Y FEDERICO 01", "PASAJE ANGELES Y FEDERICO 02", "PASAJE ANGELES Y FEDERICO 03",
    "MALILLA 05", "MALILLA 06", "MALILLA 07", "MALILLA 08", "MALILLA 14", "MALILLA 15",
    "BENICALAP 01", "BENICALAP 02", "BENICALAP 03", "BENICALAP 04", "BENICALAP 05", "BENICALAP 06"
}
case4_percent_default = 0.20

case5_percent_amenities = {
    "HOMERO 01": (0.20, 0.00), "HOMERO 02": (0.20, 0.00),
    "CARCAIXENT 01": (0.20, 8.60), "CARCAIXENT 02": (0.20, 8.60),
}
case5_props = set(case5_percent_amenities.keys())

# ----------------------
# Procesadores por caso
# ----------------------
def normalize_columns(df):
    rename_map = {
        "Nombre alojamiento": "Alojamiento",
        "Fecha entrada": "Fecha entrada",
        "Fecha salida": "Fecha salida",
        "noches": "Noches ocupadas",
        "Alquiler con tasas": "Ingreso alojamiento",
        "Extras con tasas": "Ingreso limpieza",
        "Total reserva con tasas": "Total ingresos",
        "Web origen": "Portal",
        "Comisión Portal/Intermediario: Comisión calculada": "Comisión portal",
        "IVA del alojamiento": "IVA del alquiler",
        "Portal": "Portal"
    }
    out = df.copy()
    for k, v in rename_map.items():
        if k in out.columns:
            out.rename(columns={k: v}, inplace=True)
    for c in ["Ingreso alojamiento", "Ingreso limpieza", "Total ingresos", "Comisión portal", "IVA del alquiler"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)
    return out

def apply_booking_vat_on_commission(df, portal_col="Portal", commission_col="Comisión portal"):
    out = df.copy()
    if portal_col in out.columns and commission_col in out.columns:
        mask = out[portal_col].astype(str).str.lower().str.contains("booking", na=False)
        out.loc[mask, commission_col] = out.loc[mask, commission_col] * 1.21
    return out

def process_case1(df):
    df = normalize_columns(df)
    df = apply_booking_vat_on_commission(df, "Portal", "Comisión portal")
    def honorarios(row):
        key = str(row["Alojamiento"]).strip().upper()
        pct = case1_percent_amenities.get(key, (0.20, 0.0))[0]
        return row["Ingreso alojamiento"] * pct * 1.21
    def amenities(row):
        key = str(row["Alojamiento"]).strip().upper()
        return case1_percent_amenities.get(key, (0.20, 0.0))[1]
    out = df.copy()
    out["Honorarios Florit"] = out.apply(honorarios, axis=1)
    out["Gasto limpieza"] = out["Ingreso limpieza"]
    out["Amenities"] = out.apply(amenities, axis=1)
    out["Total Gastos"] = out[["Comisión portal", "Honorarios Florit", "Gasto limpieza", "Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"] = out["Total ingresos"] - out["Comisión portal"]
    out = keep_original_dates(out, ["Fecha entrada", "Fecha salida"])
    cols = ["Alojamiento", "Fecha entrada", "Fecha salida", "Noches ocupadas",
            "Ingreso alojamiento", "Ingreso limpieza", "Total ingresos",
            "Portal", "Comisión portal", "Honorarios Florit", "Gasto limpieza",
            "Amenities", "Total Gastos", "Pago al propietario", "Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    out = out[cols].sort_values(by=["Alojamiento", "Fecha entrada"], ascending=[True, True])
    return out

def process_case2(df):
    df = normalize_columns(df)
    out = df.copy()
    if "Alojamiento" in out.columns and "Portal" in out.columns and "Comisión portal" in out.columns:
        mask_apolo = out["Alojamiento"].astype(str).str.upper().isin({"APOLO 29","APOLO 197"})
        mask_book = out["Portal"].astype(str).str.lower().str.contains("booking", na=False)
        out.loc[mask_apolo & mask_book, "Comisión portal"] = out.loc[mask_apolo & mask_book, "Comisión portal"] * 1.21
    def honorarios(row):
        key = str(row["Alojamiento"]).strip().upper()
        pct = case2_percent_amenities.get(key, (0.20, 0.0))[0]
        base = row["Ingreso alojamiento"] - row.get("IVA del alquiler", 0.0)
        return base * pct * 1.21
    def amenities(row):
        key = str(row["Alojamiento"]).strip().upper()
        return case2_percent_amenities.get(key, (0.20, 0.0))[1]
    out["Honorarios Florit"] = out.apply(honorarios, axis=1)
    out["Gasto limpieza"] = out.get("Ingreso limpieza", 0.0)
    out["Amenities"] = out.apply(amenities, axis=1)
    out["Total Gastos"] = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"] = out["Total ingresos"] - out["Comisión portal"]
    out = keep_original_dates(out, ["Fecha entrada", "Fecha salida"])
    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","IVA del alquiler","Ingreso limpieza",
            "Total ingresos","Portal","Comisión portal","Honorarios Florit","Gasto limpieza","Amenities","Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].sort_values(by=["Alojamiento","Fecha entrada"])

def process_case3(df):
    df = normalize_columns(df)
    out = apply_booking_vat_on_commission(df, "Portal", "Comisión portal")
    def honorarios(row):
        key = str(row["Alojamiento"]).strip().upper()
        pct = case3_cleaning_amenities.get(key, (0.20, None, None))[0]
        base = row["Ingreso alojamiento"] - row["Comisión portal"]
        return base * pct * 1.21
    def gasto_limpieza(row):
        key = str(row["Alojamiento"]).strip().upper()
        return case3_cleaning_amenities.get(key, (0.20, 0.0, 0.0))[1]
    def amenities(row):
        key = str(row["Alojamiento"]).strip().upper()
        return case3_cleaning_amenities.get(key, (0.20, 0.0, 0.0))[2]
    out["Honorarios Florit"] = out.apply(honorarios, axis=1)
    out["Gasto limpieza"] = out.apply(gasto_limpieza, axis=1)
    out["Amenities"] = out.apply(amenities, axis=1)
    out["Total Gastos"] = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"] = out["Total ingresos"] - out["Comisión portal"]
    out = keep_original_dates(out, ["Fecha entrada", "Fecha salida"])
    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","Ingreso limpieza","Total ingresos",
            "Portal","Comisión portal","Honorarios Florit","Gasto limpieza","Amenities","Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].sort_values(by=["Alojamiento","Fecha entrada"])

def process_case4(df):
    df = normalize_columns(df)
    out = df.copy()
    def honorarios(row):
        base = row["Ingreso alojamiento"] - row.get("IVA del alquiler", 0.0) - row["Comisión portal"]
        return base * 0.20
    out["Honorarios Florit"] = out.apply(honorarios, axis=1)
    out["Gasto limpieza"] = out.get("Ingreso limpieza", 0.0)
    out["Amenities"] = 0.0
    out["Total Gastos"] = out[["Comisión portal","Honorarios Florit"]].sum(axis=1) + out["Gasto limpieza"] + out["Amenities"]
    out["Pago al propietario"] = out["Ingreso alojamiento"] - out.get("IVA del alquiler", 0.0) - out["Comisión portal"] - out["Honorarios Florit"]
    out["Pago recibido"] = out["Ingreso alojamiento"] + out.get("Ingreso limpieza", 0.0) - out["Comisión portal"]
    out = keep_original_dates(out, ["Fecha entrada", "Fecha salida"])
    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","IVA del alquiler",
            "Portal","Comisión portal","Honorarios Florit","Pago al propietario","Pago recibido","Ingreso limpieza"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].sort_values(by=["Alojamiento","Fecha entrada"])

def process_case5(df):
    df = normalize_columns(df)
    out = df.copy()
    def honorarios(row):
        key = str(row["Alojamiento"]).strip().upper()
        pct = case5_percent_amenities.get(key, (0.20, 0.0))[0]
        base = row["Ingreso alojamiento"] - row.get("IVA del alquiler", 0.0) - row["Comisión portal"]
        return base * pct * 1.21
    def amenities(row):
        key = str(row["Alojamiento"]).strip().upper()
        return case5_percent_amenities.get(key, (0.20, 0.0))[1]
    out["Honorarios Florit"] = out.apply(honorarios, axis=1)
    out["Gasto limpieza"] = out.get("Ingreso limpieza", 0.0)
    out["Amenities"] = out.apply(amenities, axis=1)
    out["Total Gastos"] = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"] = out["Total ingresos"] - out["Comisión portal"]
    out = keep_original_dates(out, ["Fecha entrada", "Fecha salida"])
    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","IVA del alquiler",
            "Ingreso limpieza","Total ingresos","Portal","Comisión portal","Honorarios Florit","Gasto limpieza",
            "Amenities","Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].sort_values(by=["Alojamiento","Fecha entrada"])

# ----------------------
# Detección automática
# ----------------------
def detect_case(df):
    alos = set(df.get("Nombre alojamiento", df.get("Alojamiento", pd.Series(dtype=str))).astype(str).str.upper().unique())
    scores = {
        1: len(alos & set(case1_percent_amenities.keys())),
        2: len(alos & set(case2_percent_amenities.keys())),
        3: len(alos & set(case3_cleaning_amenities.keys())),
        4: len(alos & set(case4_props)),
        5: len(alos & set(case5_percent_amenities.keys())),
    }
    best_case = max(scores, key=scores.get)
    return best_case, scores

processors = {
    1: process_case1,
    2: process_case2,
    3: process_case3,
    4: process_case4,
    5: process_case5
}

# ----------------------
# UI
# ----------------------
st.title("📊 LIQUIDACIONES Automáticas (Casos 1–5)")
st.caption("Sube un Excel de Avantio (.xlsx). Detectamos el caso según alojamientos y aplicamos las reglas oficiales.")

file = st.file_uploader("Sube el archivo de reservas (.xlsx)", type=["xlsx"])

if file is not None:
    # Leer (encabezado suele estar en fila 1 = header=1)
    try:
        df = pd.read_excel(file, header=1)
    except Exception:
        df = pd.read_excel(file)

    st.write("Vista previa:")
    st.dataframe(df.head(12), use_container_width=True)

    detected_case, scores = detect_case(df)
    st.info(f"**Caso detectado automáticamente: {detected_case}** | Recuento por caso: {scores}")
    case_choice = st.selectbox("Si lo prefieres, elige manualmente el caso", options=[1,2,3,4,5], index=[1,2,3,4,5].index(detected_case))

    if st.button("Generar liquidación"):
        processor = processors[case_choice]
        result_df = processor(df)

        st.success(f"Liquidación generada (Caso {case_choice}).")
        st.dataframe(result_df.head(50), use_container_width=True)

        # Excel listo para descargar
        num_cols = [c for c in result_df.columns if pd.api.types.is_numeric_dtype(result_df[c])]
        excel_buffer = to_excel_grouped(result_df, group_col="Alojamiento", name=f"CASO {case_choice}", number_cols=num_cols)
        st.download_button(
            label="📥 Descargar Excel formateado",
            data=excel_buffer,
            file_name=f"Liquidacion_CASO_{case_choice}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.info("Sube un archivo para comenzar.")
