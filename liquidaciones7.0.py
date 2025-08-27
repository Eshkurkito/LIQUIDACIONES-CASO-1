import streamlit as st
import pandas as pd
import numpy as np
from datetime import date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import re

st.set_page_config(page_title="LIQUIDACIONES (Casos 1–5) + Export Excel", page_icon="📊", layout="wide")

# ========= Utilidades =========
def ensure_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Garantiza nombres de columna únicos sin eliminar ninguna. Añade sufijos .1, .2, ..."""
    counts = {}
    new_cols = []
    for c in df.columns:
        name = str(c)
        n = counts.get(name, 0)
        new_cols.append(name if n == 0 else f"{name}.{n}")
        counts[name] = n + 1
    out = df.copy()
    out.columns = new_cols
    return out

def _first_existing(df, candidates):
    norm_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        k = str(cand).strip().lower()
        if k in norm_map:
            return norm_map[k]
    return None

def ensure_required(df, required, ctx=""):
    missing = [c for c in required if c not in df.columns]
    if missing:
        st.error(f"Faltan columnas requeridas: {missing} en {ctx}. Ajusta el archivo o usa el modo por letras.")
        st.stop()

def fmt_es_num(x):
    try:
        return f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return x

def show_table_es(df, title):
    st.subheader(title)
    view = ensure_unique_columns(df)  # blindaje anti-duplicadas
    view = view.copy()
    for c in view.columns:
        if pd.api.types.is_numeric_dtype(view[c]):
            view[c] = view[c].apply(fmt_es_num)
    st.dataframe(view, use_container_width=True)

def find_col(df: pd.DataFrame, base: str):
    """Devuelve el nombre real de la columna que coincide EXACTA con `base` o con sufijo .1, .2, ..."""
    cols = list(df.columns)
    for c in cols:
        if str(c).strip().lower() == base.strip().lower():
            return c
    pattern = re.compile(rf"^{re.escape(base)}(\.\d+)?$", re.IGNORECASE)
    for c in cols:
        if pattern.match(str(c)):
            return c
    return None

# ========= Normalización =========
LETTER_MAP_DEFAULT = {
    "W": "Alojamiento",
    "D": "Fecha entrada",
    "F": "Fecha salida",
    "H": "Noches ocupadas",
    "I": "Ingreso alojamiento",
    "J": "Ingreso limpieza",
    "O": "Total ingresos",
    "AP": "Portal",
    "AR": "Comisión portal",
    "AL": "IVA del alquiler",
}

def letters_to_idx(letter):
    s = letter.upper()
    n = 0
    for ch in s:
        if not ('A' <= ch <= 'Z'): return None
        n = n*26 + (ord(ch)-ord('A')+1)
    return n-1

def normalize_columns_by_letters(df, letter_map=LETTER_MAP_DEFAULT):
    out = df.copy()
    cols = list(out.columns)
    rename = {}
    for L, std in letter_map.items():
        i = letters_to_idx(L)
        if i is not None and i < len(cols):
            rename[cols[i]] = std
    out.rename(columns=rename, inplace=True)
    return normalize_columns(out)

def normalize_columns(df):
    out = df.copy()

    col_aloj = _first_existing(out, ["Nombre alojamiento","Alojamiento","Nombre del alojamiento","Nombre Alojamiento"])
    col_fent = _first_existing(out, ["Fecha entrada","Fecha de entrada"])
    col_fsal = _first_existing(out, ["Fecha salida","Fecha de salida"])
    col_noch = _first_existing(out, ["Noches","noches","Noches ocupadas"])
    col_alq  = _first_existing(out, ["Alquiler con tasas","Ingreso alojamiento","Importe alojamiento"])
    col_ext  = _first_existing(out, ["Extras con tasas","Ingreso limpieza","Limpieza","Importe limpieza"])
    col_tot  = _first_existing(out, ["Total reserva con tasas","Total ingresos","Total"])
    col_port = _first_existing(out, ["Web origen","Portal","Canal","Fuente"])
    col_comi = _first_existing(out, ["Comisión Portal/Intermediario: Comisión calculada","Comisión portal","Comisión"])
    col_ivaal= _first_existing(out, ["IVA del alojamiento","IVA alojamiento","IVA del alquiler"])

    rename = {}
    if col_aloj: rename[col_aloj] = "Alojamiento"
    if col_fent: rename[col_fent] = "Fecha entrada"
    if col_fsal: rename[col_fsal] = "Fecha salida"
    if col_noch: rename[col_noch] = "Noches ocupadas"
    if col_alq:  rename[col_alq]  = "Ingreso alojamiento"
    if col_ext:  rename[col_ext]  = "Ingreso limpieza"
    if col_tot:  rename[col_tot]  = "Total ingresos"
    if col_port: rename[col_port] = "Portal"
    if col_comi: rename[col_comi] = "Comisión portal"
    if col_ivaal:rename[col_ivaal]= "IVA del alquiler"

    out.rename(columns=rename, inplace=True)

    # Tipado
    for c in ["Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisión portal","IVA del alquiler","Noches ocupadas"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)

    for c in ["Fecha entrada","Fecha salida"]:
        if c in out.columns:
            out[c] = pd.to_datetime(out[c], errors="coerce", dayfirst=True)

    if "Alojamiento" in out.columns:
        out["Alojamiento"] = out["Alojamiento"].astype(str).str.strip().str.upper()

    return out

# ========= Reglas de casos =========
# Caso 1
case1_percent_amenities = {
    "APOLO 180": (0.20, 12.04), "ALMIRANTE 01": (0.22, 11.33), "ALMIRANTE 02": (0.22, 11.33),
    "CADIZ": (0.20, 9.11), "DENIA 61": (0.20, 10.96), "DOLORES ALCAYDE 04": (0.20, 11.33),
    "DR.LLUCH": (0.20, 11.16), "ERUDITO": (0.20, 13.37), "GOZALBO": (0.20, 15.25),
    "LA ELIANA": (0.20, 15.25), "MORAIRA": (0.25, 11.33), "NAPOLES Y SICILIA": (0.25, 0.00),
    "OLIVERETA 5": (0.20, 0.00), "OVE 01": (0.18, 0.00), "OVE 02": (0.18, 0.00),
    "QUART I": (0.20, 9.09), "QUART II": (0.20, 9.09), "SAN LUIS": (0.20, 11.02),
    "SERRANOS": (0.20, 13.37), "SEVILLA": (0.18, 9.45), "TUNDIDORES": (0.20, 7.85),
    "VALLE": (0.20, 11.33),
}
case1_props = set(case1_percent_amenities.keys())

# Caso 2
case2_percent_amenities = {
    "VISITACION": (0.20, 14.88),
    "PADRE PORTA 6": (0.20, 12.09), "PADRE PORTA 7": (0.20, 12.09), "PADRE PORTA 8": (0.20, 12.09),
    "PADRE PORTA 9": (0.20, 12.09), "PADRE PORTA 10": (0.20, 12.09),
    "LLADRO Y MALLI 00": (0.20, 9.45), "LLADRO Y MALLI 01": (0.20, 9.45), "LLADRO Y MALLI 02": (0.20, 9.45),
    "LLADRO Y MALLI 03": (0.20, 9.45), "LLADRO Y MALLI 04": (0.20, 9.45),
    "APOLO 29": (0.20, 11.58), "APOLO 197": (0.20, 17.40),
}
case2_props = set(case2_percent_amenities.keys())

# Caso 3
case3_cleaning_amenities = {
    "ZAPATEROS 10-2": (0.20, 60.00, 15.24),
    "ZAPATEROS 10-6": (0.20, 75.00, 15.24),
    "ZAPATEROS 10-8": (0.20, 75.00, 15.24),
    "ZAPATEROS 12-5": (0.20, 60.00, 11.33),
    "ALFARO": (0.20, 80.00, 14.88),
}
case3_props = set(case3_cleaning_amenities.keys())

# Caso 4
case4_props = {
    "SERRERIA 04", "SERRERIA 05", "RETOR A", "RETOR B",
    "PASAJE ANGELES Y FEDERICO 01", "PASAJE ANGELES Y FEDERICO 02", "PASAJE ANGELES Y FEDERICO 03",
    "MALILLA 05", "MALILLA 06", "MALILLA 07", "MALILLA 08", "MALILLA 14", "MALILLA 15",
    "BENICALAP 01", "BENICALAP 02", "BENICALAP 03", "BENICALAP 04", "BENICALAP 05", "BENICALAP 06"
}

# Caso 5
case5_percent_amenities = {
    "HOMERO 01": (0.20, 0.00), "HOMERO 02": (0.20, 0.00),
    "CARCAIXENT 01": (0.20, 8.60), "CARCAIXENT 02": (0.20, 8.60)
}
case5_props = set(case5_percent_amenities.keys())

def props_for_case(case):
    if case == 1: return case1_props
    if case == 2: return case2_props
    if case == 3: return case3_props
    if case == 4: return case4_props
    if case == 5: return case5_props
    return set()

# ========= Reglas transversales =========
def adjust_commission_booking_and_empty_portal(df, treat_empty_as_booking=False):
    """
    Aplica +21% a comisión si Portal contiene 'booking'. Si treat_empty_as_booking=True,
    también aplica +21% a filas con Portal vacío. Devuelve (df_ajustado, num_warn).
    """
    out = df.copy()
    portal_col = "Portal"
    commission_col = "Comisión portal"
    if portal_col not in out.columns or commission_col not in out.columns:
        return out, 0

    ser = out[portal_col]
    if isinstance(ser, pd.DataFrame):
        ser = ser.iloc[:, 0]
    ser = ser.astype("string").fillna("")

    out[commission_col] = pd.to_numeric(out[commission_col], errors="coerce").fillna(0.0)

    mask_booking = ser.str.lower().str.contains("booking", na=False)
    mask_empty   = ser.str.strip().eq("")
    warn_count = int(((mask_empty) & (out[commission_col] > 0)).sum())

    out.loc[mask_booking, commission_col] = out.loc[mask_booking, commission_col] * 1.21
    if treat_empty_as_booking:
        out.loc[mask_empty, commission_col] = out.loc[mask_empty, commission_col] * 1.21

    return out, warn_count

# ========= Procesadores (SIN prorrateo; se filtra por Fecha de entrada) =========
def process_case1(df, treat_empty_as_booking=False):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisión portal","Portal"], "Caso 1")
    df, warn_count = adjust_commission_booking_and_empty_portal(df, treat_empty_as_booking)

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = case1_percent_amenities.get(key,(0.20,0.0))[0]
        return float(r.get("Ingreso alojamiento",0.0)) * pct * 1.21

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case1_percent_amenities.get(key,(0.20,0.0))[1])

    out = df.copy()
    out["Honorarios Florit"] = out.apply(honorarios, axis=1)
    out["Gasto limpieza"]   = out.get("Ingreso limpieza", 0.0)
    out["Amenities"]        = out.apply(amenities, axis=1)
    out["Total Gastos"]     = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"]    = out["Total ingresos"] - out["Comisión portal"]

    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","Ingreso limpieza",
            "Total ingresos","Portal","Comisión portal","Honorarios Florit","Gasto limpieza","Amenities",
            "Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

def process_case2(df, treat_empty_as_booking=False):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Total ingresos","Comisión portal","Portal"], "Caso 2")
    df, warn_count = adjust_commission_booking_and_empty_portal(df, treat_empty_as_booking)

    # +21% adicional en Booking para APOLO 29/197 (si aplica)
    mask_apolo = df["Alojamiento"].astype(str).str.upper().isin({"APOLO 29","APOLO 197"})
    mask_book  = df["Portal"].astype(str).str.lower().str.contains("booking", na=False)
    df.loc[mask_apolo & mask_book, "Comisión portal"] = pd.to_numeric(df.loc[mask_apolo & mask_book, "Comisión portal"], errors="coerce").fillna(0.0) * 1.21

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = case2_percent_amenities.get(key,(0.20,0.0))[0]
        base = float(r.get("Ingreso alojamiento",0.0)) - float(r.get("IVA del alquiler",0.0))
        return base * pct * 1.21

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case2_percent_amenities.get(key,(0.20,0.0))[1])

    out = df.copy()
    out["Honorarios Florit"] = out.apply(honorarios, axis=1)
    out["Gasto limpieza"]   = out.get("Ingreso limpieza", 0.0)
    out["Amenities"]        = out.apply(amenities, axis=1)
    out["Total Gastos"]     = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"]    = out["Total ingresos"] - out["Comisión portal"]

    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","IVA del alquiler",
            "Ingreso limpieza","Total ingresos","Portal","Comisión portal","Honorarios Florit","Gasto limpieza",
            "Amenities","Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

def process_case3(df, treat_empty_as_booking=False):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Total ingresos","Comisión portal","Portal"], "Caso 3")
    df, warn_count = adjust_commission_booking_and_empty_portal(df, treat_empty_as_booking)

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = case3_cleaning_amenities.get(key,(0.20,0.0,0.0))[0]
        base = float(r.get("Ingreso alojamiento",0.0)) - float(r.get("Comisión portal",0.0))
        return base * pct * 1.21

    def gasto_limpieza(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case3_cleaning_amenities.get(key,(0.20,0.0,0.0))[1])

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case3_cleaning_amenities.get(key,(0.20,0.0,0.0))[2])

    out = df.copy()
    out["Honorarios Florit"] = out.apply(honorarios, axis=1)
    out["Gasto limpieza"]   = out.apply(gasto_limpieza, axis=1)
    out["Amenities"]        = out.apply(amenities, axis=1)
    out["Total Gastos"]     = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"]    = out["Total ingresos"] - out["Comisión portal"]

    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","Ingreso limpieza",
            "Total ingresos","Portal","Comisión portal","Honorarios Florit","Gasto limpieza","Amenities",
            "Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

def process_case4(df, treat_empty_as_booking=False):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Comisión portal"], "Caso 4")
    df["Portal"] = df.get("Portal", pd.Series([""]*len(df)))
    df, warn_count = adjust_commission_booking_and_empty_portal(df, treat_empty_as_booking)

    out = df.copy()

    # IVA del alquiler = 10% del ingreso alojamiento
    out["IVA del alquiler"] = pd.to_numeric(out.get("Ingreso alojamiento", 0.0), errors="coerce").fillna(0.0) * 0.10

    # Honorarios = (Ingreso - IVA - Comisión) * 20%
    def honorarios(r):
        base = float(r.get("Ingreso alojamiento",0.0)) - float(r.get("IVA del alquiler",0.0)) - float(r.get("Comisión portal",0.0))
        return base * 0.20

    out["Honorarios Florit"] = out.apply(honorarios, axis=1)

    # Pago al propietario
    out["Pago al propietario"] = (
        pd.to_numeric(out.get("Ingreso alojamiento",0.0), errors="coerce").fillna(0.0)
        - pd.to_numeric(out.get("IVA del alquiler",0.0), errors="coerce").fillna(0.0)
        - pd.to_numeric(out.get("Comisión portal",0.0), errors="coerce").fillna(0.0)
        - pd.to_numeric(out.get("Honorarios Florit",0.0), errors="coerce").fillna(0.0)
    ).round(2)

    # Pago recibido
    out["Pago recibido"] = (
        pd.to_numeric(out.get("Ingreso alojamiento",0.0), errors="coerce").fillna(0.0)
        + pd.to_numeric(out.get("Ingreso limpieza",0.0), errors="coerce").fillna(0.0)
        - pd.to_numeric(out.get("Comisión portal",0.0), errors="coerce").fillna(0.0)
    ).round(2)

    # columnas de salida: (sin Gasto limpieza, Amenities, Total Gastos)
    cols = [
        "Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas",
        "Ingreso alojamiento","IVA del alquiler","Ingreso limpieza","Total ingresos",
        "Portal","Comisión portal","Honorarios Florit","Pago al propietario","Pago recibido"
    ]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

def process_case5(df, treat_empty_as_booking=False):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Total ingresos","Comisión portal"], "Caso 5")
    df["Portal"] = df.get("Portal", pd.Series([""]*len(df)))
    df, warn_count = adjust_commission_booking_and_empty_portal(df, treat_empty_as_booking)

    out = df.copy()

    # IVA del alquiler = 10% del ingreso alojamiento
    out["IVA del alquiler"] = pd.to_numeric(out.get("Ingreso alojamiento", 0.0), errors="coerce").fillna(0.0) * 0.10

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = case5_percent_amenities.get(key,(0.20,0.0))[0]
        base = float(r.get("Ingreso alojamiento",0.0)) - float(r.get("IVA del alquiler",0.0)) - float(r.get("Comisión portal",0.0))
        return base * pct * 1.21

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case5_percent_amenities.get(key,(0.20,0.0))[1])

    out["Honorarios Florit"] = out.apply(honorarios, axis=1)
    out["Gasto limpieza"]   = out.get("Ingreso limpieza", 0.0)
    out["Amenities"]        = out.apply(amenities, axis=1)
    out["Total Gastos"]     = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"]    = out["Total ingresos"] - out["Comisión portal"]

    cols = [
        "Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas",
        "Ingreso alojamiento","IVA del alquiler","Ingreso limpieza","Total ingresos","Portal","Comisión portal",
        "Honorarios Florit","Gasto limpieza","Amenities","Total Gastos","Pago al propietario","Pago recibido"
    ]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

processors = {1: process_case1, 2: process_case2, 3: process_case3, 4: process_case4, 5: process_case5}

# ========= Exportación a Excel formateado =========
BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def write_grouped_sheet(workbook, df, title="Liquidación"):
    ws = workbook.active
    ws.title = title

    cols = list(df.columns)

    def write_table(start_row, subdf):
        # cabecera
        for j, col in enumerate(cols, start=1):
            cell = ws.cell(row=start_row, column=j, value=col)
            cell.font = Font(bold=True)
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # filas
        for i, (_, row) in enumerate(subdf.iterrows(), start=1):
            for j, col in enumerate(cols, start=1):
                val = row[col]
                c = ws.cell(row=start_row+i, column=j, value=val)
                c.border = BORDER_THIN
                if isinstance(val, (int, float)) and not pd.isna(val):
                    c.number_format = "#.##0,00"
                else:
                    c.alignment = Alignment(wrap_text=True)
        # sumatorios debajo (solo numéricos)
        sum_row = start_row + len(subdf) + 1
        ws.cell(row=sum_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=sum_row, column=1).border = BORDER_THIN
        for j, col in enumerate(cols, start=1):
            if j == 1:
                continue
            if pd.api.types.is_numeric_dtype(subdf[col]):
                top = start_row+1
                bottom = start_row+len(subdf)
                formula = f"=SUM({get_column_letter(j)}{top}:{get_column_letter(j)}{bottom})"
                c = ws.cell(row=sum_row, column=j, value=formula)
                c.font = Font(bold=True)
                c.border = BORDER_THIN
                c.number_format = "#.##0,00"
            else:
                c = ws.cell(row=sum_row, column=j, value="")
                c.border = BORDER_THIN
        return sum_row + 2

    current_row = 1
    if "Alojamiento" in df.columns:
        for aloj, subdf in df.groupby("Alojamiento"):
            ws.cell(row=current_row, column=1, value=str(aloj)).font = Font(bold=True, size=12)
            current_row += 1
            current_row = write_table(current_row, subdf)
    else:
        current_row = write_table(current_row, df)

    for j, col in enumerate(cols, start=1):
        max_len = len(str(col))
        for r in range(1, ws.max_row+1):
            v = ws.cell(row=r, column=j).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(j)].width = min(max_len+2, 45)

def build_excel_file(df_final, filename="Liquidacion.xlsx"):
    wb = Workbook()
    write_grouped_sheet(wb, df_final, title="Liquidación")
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    st.download_button(
        label="📥 Descargar Excel (Liquidación)",
        data=bio.getvalue(),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ========= UI =========
st.title("📊 LIQUIDACIONES Automáticas (Casos 1–5) + Export Excel")
st.caption("Sube el Excel, elige período y caso, y pulsa Generar. La imputación es por Fecha de entrada (sin prorrateo).")

with st.sidebar:
    st.header("Parámetros")
    c1, c2 = st.columns(2)
    with c1:
        start_date = st.date_input("Desde", value=date(date.today().year, date.today().month, 1))
    with c2:
        end_date   = st.date_input("Hasta",  value=date(date.today().year, date.today().month, 28))
    st.divider()
    case_choice = st.radio("Caso", [1,2,3,4,5], horizontal=True)
    st.checkbox("Lectura por letras (fallback)", value=False, key="by_letters")
    st.caption("Mapeo: W, D, F, H, I, J, O, AP, AR, AL (si Avantio cambia títulos).")
    st.divider()
    treat_empty_as_booking = st.checkbox("Tratar reservas sin portal como Booking (+21% comisión)", value=False)
    generate = st.button("Generar liquidación")

file = st.file_uploader("Sube el archivo de reservas (.xlsx)", type=["xlsx"])

# ========= Generación =========
if generate:
    if not file:
        st.error("Sube primero el archivo de reservas (.xlsx).")
        st.stop()

    df_in = pd.read_excel(file, header=0)
    df_in = ensure_unique_columns(df_in)

    df_norm = normalize_columns_by_letters(df_in) if st.session_state.by_letters else normalize_columns(df_in)
    df_norm = ensure_unique_columns(df_norm)

    # Filtrar por propiedades del caso (si las hay)
    props = props_for_case(case_choice)
    if props and "Alojamiento" in df_norm.columns:
        df_norm = df_norm[df_norm["Alojamiento"].isin(props)]

    # Filtro por periodo: FECHA DE ENTRADA (SIN PRORRATEO)
    if "Fecha entrada" in df_norm.columns:
        mask = (df_norm["Fecha entrada"] >= pd.to_datetime(start_date)) & (df_norm["Fecha entrada"] <= pd.to_datetime(end_date))
        df_norm = df_norm[mask]

    # Procesar caso (ya sin prorrateos)
    base_df, warn_count = processors[case_choice](df_norm.copy(), treat_empty_as_booking=treat_empty_as_booking)
    base_df = ensure_unique_columns(base_df)

    if warn_count > 0 and not treat_empty_as_booking:
        st.warning(f"Hay {warn_count} reservas con comisión > 0 pero portal vacío. "
                   f"Si deben ser Booking, marca «Tratar reservas sin portal como Booking (+21% comisión)» y vuelve a generar.")

    st.success(f"Liquidación generada (Caso {case_choice}) • {start_date.strftime('%d/%m/%Y')}–{end_date.strftime('%d/%m/%Y')}")
    # Orden final: por Alojamiento y Fecha entrada si existe
    sort_cols = [c for c in ["Alojamiento","Fecha entrada"] if c in base_df.columns]
    if sort_cols:
        base_df = base_df.sort_values(by=sort_cols)
    show_table_es(base_df, "Tabla de liquidaciones")

    # Resumen por alojamiento
    aloj_col = find_col(base_df, "Alojamiento")
    pago_col = find_col(base_df, "Pago al propietario")
    if aloj_col is not None and pago_col is not None:
        pagos = (base_df[[aloj_col, pago_col]]
                 .groupby(aloj_col, as_index=False)[pago_col]
                 .sum().round(2).sort_values(aloj_col))
        pagos.rename(columns={aloj_col: "Alojamiento", pago_col: "Pago al propietario"}, inplace=True)
        show_table_es(pagos, "💸 Pagos por alojamiento (suma)")
        total_general = pd.to_numeric(pagos["Pago al propietario"], errors="coerce").fillna(0.0).sum()
        st.markdown(f"**Total general a transferir:** {fmt_es_num(total_general)} €")

    # Export Excel
    file_case_name = f"Liquidacion_CASO{case_choice}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    build_excel_file(base_df, filename=file_case_name)
