#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
liquidaciones_con_banco_v3.py

Genera liquidaciones por caso y concilia con extracto bancario.
Reglas solicitadas:

- Aplicar +21% (IVA) sobre "Comisión portal" SOLO para reservas donde el Portal es Booking,
  y según estos criterios por caso:
    * Caso 1 -> TODOS los alojamientos
    * Caso 2 -> SOLO "APOLO 197" y "APOLO 29"
    * Caso 3 -> TODOS los alojamientos
    * Caso 4 -> NINGUNO
    * Caso 5 -> NINGUNO

- La conciliación usa SIEMPRE la columna "Pago recibido", no "Pago al propietario".
- Permite elegir fila de cabecera para el Excel de liquidaciones.
- Permite indicar que la columna del portal viene en la letra "AP".
- Permite fijar una ventana de días para la fecha bancaria, o sin límite (por defecto).

Salida:
- Excel con liquidaciones por caso
- Excel con conciliación: "Conciliados", "Banco_sin_match", "Liq_sin_match"

Ejemplo de uso:
    python liquidaciones_con_banco_v3.py \
        --liq "Liquidaciones max 3.xlsx" --liq-header 2 --portal-col AP \
        --bank "BBVA Histórico movimientos.xlsx" --bank-sheet Historico --bank-header 14 \
        --filter-by entry --month 7 --year 2025 \
        --date-window-days -1 \
        --out-liq "Liquidaciones_por_caso_Julio2025.xlsx" \
        --out-recon "Conciliacion_Banco_vs_Liquidaciones_Julio2025.xlsx"
"""

from __future__ import annotations
import argparse
from typing import Optional, List, Set, Tuple
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
from pathlib import Path

# ============================= Utilidades base =============================

def ensure_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    counts, new_cols = {}, []
    for c in df.columns:
        name = str(c)
        n = counts.get(name, 0)
        new_cols.append(name if n == 0 else f"{name}.{n}")
        counts[name] = n + 1
    out = df.copy()
    out.columns = new_cols
    return out

def _first_existing(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    norm_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        k = str(cand).strip().lower()
        if k in norm_map:
            return norm_map[k]
    return None

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    # Intentos por nombre habitual
    col_aloj = _first_existing(out, ["Nombre alojamiento","Alojamiento","Nombre del alojamiento","Nombre Alojamiento"])
    col_fent = _first_existing(out, ["Fecha entrada","Fecha de entrada"])
    col_fsal = _first_existing(out, ["Fecha salida","Fecha de salida"])
    col_noch = _first_existing(out, ["Noches","noches","Noches ocupadas"])
    col_alq  = _first_existing(out, ["Alquiler con tasas","Ingreso alojamiento","Importe alojamiento"])
    col_ext  = _first_existing(out, [
        "Ingreso limpieza","Tarifa limpieza","Limpieza","Importe limpieza",
        "Extras con tasas","Gastos de limpieza","Gasto limpieza"
    ])
    col_tot  = _first_existing(out, ["Total reserva con tasas","Total ingresos","Total"])
    col_port = _first_existing(out, ["Portal","Web origen","Canal","Fuente"])
    col_comi = _first_existing(out, ["Comisión Portal/Intermediario: Comisión calculada","Comisión portal","Comisión"])
    col_ivaal= _first_existing(out, ["IVA del alojamiento","IVA alojamiento","IVA del alquiler"])

    rename = {}
    if col_aloj: rename[col_aloj] = "Alojamiento"
    if col_fent: rename[col_fent] = "Fecha entrada"
    if col_fsal: rename[col_fsal] = "Fecha salida"
    if col_noch: rename[col_noch] = "Noches ocupadas"
    if col_alq:  rename[col_alq]  = "Ingreso alojamiento"
    if col_ext:  rename[col_ext]  = "Ingreso limpieza"
    if col_tot:  rename[col_tot]  = "Total ingresos"
    if col_port: rename[col_port] = "Portal"
    if col_comi: rename[col_comi] = "Comisión portal"
    if col_ivaal:rename[col_ivaal]= "IVA del alquiler"

    out.rename(columns=rename, inplace=True)

    # Tipado
    for c in ["Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisión portal","IVA del alquiler","Noches ocupadas"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)

    for c in ["Fecha entrada","Fecha salida"]:
        if c in out.columns:
            out[c] = pd.to_datetime(out[c], errors="coerce", dayfirst=True)

    if "Alojamiento" in out.columns:
        out["Alojamiento"] = out["Alojamiento"].astype(str).str.strip().str.upper()
    if "Portal" in out.columns:
        out["Portal"] = out["Portal"].astype(str).str.strip()
    if "Noches ocupadas" in out.columns:
        out["Noches ocupadas"] = pd.to_numeric(out["Noches ocupadas"], errors="coerce").fillna(0).round(0).astype(int)

    return out

def num_series(df: pd.DataFrame, col: str, default: float=0.0) -> pd.Series:
    if col in df.columns:
        return pd.to_numeric(df[col], errors="coerce").fillna(default)
    else:
        return pd.Series([default]*len(df), index=df.index, dtype="float")

# ========================= Reglas de casos & amenities =========================

CASE1_AMENITIES = {
    "APOLO 180": (0.20, 12.04), "ALMIRANTE 01": (0.22, 11.33), "ALMIRANTE 02": (0.22, 11.33),
    "CADIZ": (0.20, 9.11), "DENIA 61": (0.20, 10.96), "DOLORES ALCAYDE 04": (0.20, 11.33),
    "DR.LLUCH": (0.20, 11.16), "ERUDITO": (0.20, 13.37), "GOZALBO": (0.20, 15.25),
    "LA ELIANA": (0.20, 15.25), "MORAIRA": (0.25, 11.33), "NAPOLES Y SICILIA": (0.25, 0.00),
    "OLIVERETA 5": (0.20, 0.00), "OVE 01": (0.18, 0.00), "OVE 02": (0.18, 0.00),
    "QUART I": (0.20, 9.09), "QUART II": (0.20, 9.09), "SAN LUIS": (0.20, 11.02),
    "SERRANOS": (0.20, 13.37), "SEVILLA": (0.18, 9.45), "TUNDIDORES": (0.20, 7.85),
    "VALLE": (0.20, 11.33),
}
CASE2_AMENITIES = {
    "VISITACION": (0.20, 14.88),
    "PADRE PORTA 6": (0.20, 12.09), "PADRE PORTA 7": (0.20, 12.09), "PADRE PORTA 8": (0.20, 12.09),
    "PADRE PORTA 9": (0.20, 12.09), "PADRE PORTA 10": (0.20, 12.09),
    "LLADRO Y MALLI 00": (0.20, 9.45), "LLADRO Y MALLI 01": (0.20, 9.45), "LLADRO Y MALLI 02": (0.20, 9.45),
    "LLADRO Y MALLI 03": (0.20, 9.45), "LLADRO Y MALLI 04": (0.20, 9.45),
    "APOLO 29": (0.20, 11.58), "APOLO 197": (0.20, 17.40),
}
CASE3_CLEANING_AMEN = {
    "ZAPATEROS 10-2": (0.20, 60.00, 15.24),
    "ZAPATEROS 10-6": (0.20, 75.00, 15.24),
    "ZAPATEROS 10-8": (0.20, 75.00, 15.24),
    "ZAPATEROS 12-5": (0.20, 60.00, 11.33),
    "ALFARO": (0.20, 80.00, 14.88),
}
CASE4_PROPS = {
    "SERRERIA 04", "SERRERIA 05", "RETOR A", "RETOR B",
    "PASAJE ANGELES Y FEDERICO 01", "PASAJE ANGELES Y FEDERICO 02", "PASAJE ANGELES Y FEDERICO 03",
    "MALILLA 05", "MALILLA 06", "MALILLA 07", "MALILLA 08", "MALILLA 14", "MALILLA 15",
    "BENICALAP 01", "BENICALAP 02", "BENICALAP 03", "BENICALAP 04", "BENICALAP 05", "BENICALAP 06"
}
CASE5_AMENITIES = {
    "HOMERO 01": (0.20, 0.00), "HOMERO 02": (0.20, 0.00),
    "CARCAIXENT 01": (0.20, 8.60), "CARCAIXENT 02": (0.20, 8.60)
}

def case_of(aloj: str):
    a = str(aloj).strip().upper()
    if a in CASE1_AMENITIES: return 1
    if a in CASE2_AMENITIES: return 2
    if a in CASE3_CLEANING_AMEN: return 3
    if a in CASE4_PROPS: return 4
    if a in CASE5_AMENITIES: return 5
    return None

def should_apply_booking_vat(case_num, alojamiento):
    if case_num is None:
        return False
    a = str(alojamiento).strip().upper()
    if case_num == 1: return True
    if case_num == 2: return a in {"APOLO 197","APOLO 29"}
    if case_num == 3: return True
    if case_num in (4,5): return False
    return False

# ============================= Procesado por casos =============================

def process_generic(df: pd.DataFrame) -> pd.DataFrame:
    df = normalize_columns(df)

    # Ajuste +21% a Comisión portal SOLO si: Portal es Booking y la regla del caso lo indica
    if "Portal" in df.columns and "Comisión portal" in df.columns:
        is_booking = df["Portal"].astype(str).str.lower().str.contains("booking", na=False)
        adj = []
        for i, r in df.iterrows():
            if not bool(is_booking.loc[i]):
                adj.append(False); continue
            cnum = case_of(r.get("Alojamiento",""))
            adj.append(should_apply_booking_vat(cnum, r.get("Alojamiento","")))
        idx = pd.Series(adj, index=df.index)
        df.loc[idx, "Comisión portal"] = pd.to_numeric(df.loc[idx, "Comisión portal"], errors="coerce").fillna(0.0) * 1.21

    # Particionadores por caso y fórmulas
    def filter_props(df0: pd.DataFrame, props_set: Set[str]) -> pd.DataFrame:
        if "Alojamiento" not in df0.columns:
            return df0.iloc[0:0].copy()
        return df0[df0["Alojamiento"].isin({p.upper() for p in props_set})].copy()

    def num_series(df, col, default=0.0):
        if col in df.columns:
            return pd.to_numeric(df[col], errors="coerce").fillna(default)
        else:
            return pd.Series([default]*len(df), index=df.index, dtype="float")

    # Caso 1
    c1 = filter_props(df, set(CASE1_AMENITIES.keys()))
    if len(c1)>0:
        def honor1(r):
            key = str(r.get("Alojamiento","")).strip().upper()
            pct = CASE1_AMENITIES.get(key,(0.20,0.0))[0]
            return float(r.get("Ingreso alojamiento",0.0)) * pct * 1.21
        def amen1(r):
            key = str(r.get("Alojamiento","")).strip().upper()
            return float(CASE1_AMENITIES.get(key,(0.20,0.0))[1])
        c1["Honorarios Florit"] = c1.apply(honor1, axis=1).round(2)
        c1["Gasto limpieza"]   = num_series(c1,"Ingreso limpieza").round(2)
        c1["Amenities"]        = c1.apply(amen1, axis=1).round(2)
        c1["Total Gastos"]     = (num_series(c1,"Comisión portal")+num_series(c1,"Honorarios Florit")+num_series(c1,"Gasto limpieza")+num_series(c1,"Amenities")).round(2)
        c1["Pago al propietario"] = (num_series(c1,"Total ingresos") - num_series(c1,"Total Gastos")).round(2)
        c1["Pago recibido"]    = (num_series(c1,"Total ingresos") - num_series(c1,"Comisión portal")).round(2)

    # Caso 2
    c2 = filter_props(df, set(CASE2_AMENITIES.keys()))
    if len(c2)>0:
        ingreso = num_series(c2,"Ingreso alojamiento")
        c2["IVA del alquiler"] = ingreso - (ingreso / 1.10)
        def honor2(r):
            key = str(r.get("Alojamiento","")).strip().upper()
            pct = CASE2_AMENITIES.get(key,(0.20,0.0))[0]
            base = float(r.get("Ingreso alojamiento",0.0)) - float(r.get("IVA del alquiler",0.0))
            return base * pct * 1.21
        def amen2(r):
            key = str(r.get("Alojamiento","")).strip().upper()
            return float(CASE2_AMENITIES.get(key,(0.20,0.0))[1])
        c2["Honorarios Florit"] = c2.apply(honor2, axis=1).round(2)
        c2["Gasto limpieza"]   = num_series(c2,"Ingreso limpieza").round(2)
        c2["Amenities"]        = c2.apply(amen2, axis=1).round(2)
        c2["Total Gastos"]     = (num_series(c2,"Comisión portal")+num_series(c2,"Honorarios Florit")+num_series(c2,"Gasto limpieza")+num_series(c2,"Amenities")).round(2)
        c2["Pago al propietario"] = (num_series(c2,"Total ingresos") - num_series(c2,"Total Gastos")).round(2)
        c2["Pago recibido"]    = (num_series(c2,"Total ingresos") - num_series(c2,"Comisión portal")).round(2)

    # Caso 3
    c3 = filter_props(df, set(CASE3_CLEANING_AMEN.keys()))
    if len(c3)>0:
        def honor3(r):
            key = str(r.get("Alojamiento","")).strip().upper()
            pct = CASE3_CLEANING_AMEN.get(key,(0.20,0.0,0.0))[0]
            base = float(r.get("Ingreso alojamiento",0.0)) - float(r.get("Comisión portal",0.0))
            return base * pct * 1.21
        def limp3(r):
            key = str(r.get("Alojamiento","")).strip().upper()
            return float(CASE3_CLEANING_AMEN.get(key,(0.20,0.0,0.0))[1])
        def amen3(r):
            key = str(r.get("Alojamiento","")).strip().upper()
            return float(CASE3_CLEANING_AMEN.get(key,(0.20,0.0,0.0))[2])
        c3["Honorarios Florit"] = c3.apply(honor3, axis=1).round(2)
        c3["Gasto limpieza"]   = c3.apply(limp3, axis=1).round(2)
        c3["Amenities"]        = c3.apply(amen3, axis=1).round(2)
        c3["Total Gastos"]     = (num_series(c3,"Comisión portal")+num_series(c3,"Honorarios Florit")+num_series(c3,"Gasto limpieza")+num_series(c3,"Amenities")).round(2)
        c3["Pago al propietario"] = (num_series(c3,"Total ingresos") - num_series(c3,"Total Gastos")).round(2)
        c3["Pago recibido"]    = (num_series(c3,"Total ingresos") - num_series(c3,"Comisión portal")).round(2)

    # Caso 4
    c4 = filter_props(df, set(CASE4_PROPS))
    if len(c4)>0:
        ingreso = num_series(c4,"Ingreso alojamiento")
        c4["IVA del alquiler"] = ingreso - (ingreso / 1.10)
        def honor4(r):
            base = float(r.get("Ingreso alojamiento",0.0)) - float(r.get("IVA del alquiler",0.0)) - float(r.get("Comisión portal",0.0))
            return base * 0.20
        c4["Honorarios Florit"] = c4.apply(honor4, axis=1).round(2)
        c4["Pago al propietario"] = (num_series(c4,"Ingreso alojamiento") - num_series(c4,"IVA del alquiler") - num_series(c4,"Comisión portal") - num_series(c4,"Honorarios Florit")).round(2)
        c4["Pago recibido"] = (num_series(c4,"Ingreso alojamiento") + num_series(c4,"Ingreso limpieza") - num_series(c4,"Comisión portal")).round(2)

    # Caso 5
    c5 = filter_props(df, set(CASE5_AMENITIES.keys()))
    if len(c5)>0:
        ingreso = num_series(c5,"Ingreso alojamiento")
        c5["IVA del alquiler"] = ingreso - (ingreso / 1.10)
        def honor5(r):
            key = str(r.get("Alojamiento","")).strip().upper()
            pct = CASE5_AMENITIES.get(key,(0.20,0.0))[0]
            base = float(r.get("Ingreso alojamiento",0.0)) - float(r.get("IVA del alquiler",0.0)) - float(r.get("Comisión portal",0.0))
            return base * pct * 1.21
        def amen5(r):
            key = str(r.get("Alojamiento","")).strip().upper()
            return float(CASE5_AMENITIES.get(key,(0.20,0.0))[1])
        c5["Honorarios Florit"] = c5.apply(honor5, axis=1).round(2)
        c5["Gasto limpieza"]   = num_series(c5,"Ingreso limpieza").round(2)
        c5["Amenities"]        = c5.apply(amen5, axis=1).round(2)
        c5["Total Gastos"]     = (num_series(c5,"Comisión portal")+num_series(c5,"Honorarios Florit")+num_series(c5,"Gasto limpieza")+num_series(c5,"Amenities")).round(2)
        c5["Pago al propietario"] = (num_series(c5,"Total ingresos") - num_series(c5,"Total Gastos")).round(2)
        c5["Pago recibido"]    = (num_series(c5,"Total ingresos") - num_series(c5,"Comisión portal")).round(2)

    out = pd.concat([c1,c2,c3,c4,c5], ignore_index=True, sort=False)
    if "Fecha entrada" in out.columns:
        out = out.sort_values(by=[c for c in ["Alojamiento","Fecha entrada"] if c in out.columns])
    for c in out.columns:
        if c != "Noches ocupadas" and pd.api.types.is_numeric_dtype(out[c]):
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0).round(2)
    return out

# ============================= Conciliación banco =============================

def reconcile_by_pago_recibido(liq: pd.DataFrame, bank: pd.DataFrame, date_window_days: Optional[int]=None) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    liq = liq.copy(); bank = bank.copy()
    liq["Pago recibido"] = pd.to_numeric(liq.get("Pago recibido", np.nan), errors="coerce").round(2)
    liq["Fecha entrada"] = pd.to_datetime(liq.get("Fecha entrada"), errors="coerce")
    liq["Fecha salida"]  = pd.to_datetime(liq.get("Fecha salida"), errors="coerce")
    ref_col = "Fecha salida" if "Fecha salida" in liq.columns else "Fecha entrada"
    liq["__FechaRef__"] = pd.to_datetime(liq[ref_col], errors="coerce")
    bank["ImporteAbs"] = pd.to_numeric(bank.get("Importe"), errors="coerce").abs().round(2)
    bank["Fecha"] = pd.to_datetime(bank.get("Fecha"), errors="coerce")

    matches = []
    used_bank_idx = set()
    for i, rf in liq.dropna(subset=["Pago recibido"]).iterrows():
        amt = round(abs(rf["Pago recibido"]), 2)
        cand = bank[(bank["ImporteAbs"] == amt) & (~bank.index.isin(used_bank_idx))]
        if date_window_days is not None and date_window_days >= 0:
            f_ref = rf["__FechaRef__"]
            if not pd.isna(f_ref):
                cand = cand[(cand["Fecha"] >= f_ref) & (cand["Fecha"] <= f_ref + pd.Timedelta(days=date_window_days))]
        if len(cand) > 0:
            rb = cand.sort_values("Fecha").iloc[0]
            used_bank_idx.add(rb.name)
            matches.append({
                "Fecha mov.": rb["Fecha"].date() if not pd.isna(rb["Fecha"]) else None,
                "Concepto": rb.get("Concepto",""),
                "Benef./Ord.": rb.get("Beneficiario/Ordenante",""),
                "Importe mov.": rb.get("Importe", np.nan),
                "Alojamiento": rf.get("Alojamiento",""),
                "Portal": rf.get("Portal",""),
                "Fecha entrada": rf.get("Fecha entrada",""),
                "Fecha salida": rf.get("Fecha salida",""),
                "Pago recibido": rf.get("Pago recibido", np.nan)
            })

    df_matches = pd.DataFrame(matches)
    unmatched_bank = bank[~bank.index.isin(used_bank_idx)].copy()
    matched_amts = set(round(abs(r["Pago recibido"]),2) for _, r in df_matches.iterrows())
    unmatched_liq = liq[~liq["Pago recibido"].round(2).isin(matched_amts)].copy()
    return df_matches, unmatched_bank, unmatched_liq

# ============================= I/O helpers =============================

def write_df(ws, dfi: pd.DataFrame):
    if dfi is None or len(dfi)==0:
        ws.append(["(sin datos)"]); return
    for j, col in enumerate(dfi.columns, start=1):
        ws.cell(row=1, column=j, value=str(col)).font = Font(bold=True)
    for i, (_, row) in enumerate(dfi.iterrows(), start=2):
        for j, col in enumerate(dfi.columns, start=1):
            val = row[col]
            ws.cell(row=i, column=j, value=(val if (isinstance(val,(int,float,str)) or val is None or pd.isna(val)) else str(val)))

def export_liquidaciones_by_case(out: pd.DataFrame, path: Path, case_maps: Tuple[dict,dict,dict,set,dict]):
    wb = Workbook()
    def filter_props(df0: pd.DataFrame, props_set: Set[str]) -> pd.DataFrame:
        if "Alojamiento" not in df0.columns:
            return df0.iloc[0:0].copy()
        return df0[df0["Alojamiento"].isin({p.upper() for p in props_set})].copy()
    tabs = [
        ("Caso 1", filter_props(out, set(case_maps[0].keys()))),
        ("Caso 2", filter_props(out, set(case_maps[1].keys()))),
        ("Caso 3", filter_props(out, set(case_maps[2].keys()))),
        ("Caso 4", filter_props(out, set(case_maps[3]))),
        ("Caso 5", filter_props(out, set(case_maps[4].keys()))),
    ]
    first=True
    for name, dfi in tabs:
        if first:
            ws = wb.active; ws.title = name; first=False
        else:
            ws = wb.create_sheet(name)
        write_df(ws, dfi)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    path.write_bytes(bio.getvalue())

def export_reconciliation(matches: pd.DataFrame, bank_un: pd.DataFrame, liq_un: pd.DataFrame, path: Path):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Conciliados"
    write_df(ws1, matches)
    ws2 = wb.create_sheet("Banco_sin_match"); write_df(ws2, bank_un)
    ws3 = wb.create_sheet("Liq_sin_match")
    cols = [c for c in ["Alojamiento","Portal","Fecha entrada","Fecha salida","Pago recibido"] if c in liq_un.columns]
    write_df(ws3, liq_un[cols] if len(cols)>0 else liq_un)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    path.write_bytes(bio.getvalue())

# ============================= Main =============================

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--liq", required=True, help="Excel de liquidaciones (input)")
    ap.add_argument("--liq-header", type=int, default=1, help="Fila de cabecera 1-based (default 1). Se restará 1 para pandas.")
    ap.add_argument("--portal-col", type=str, default="AP", help="Letra/nombre de la columna Portal en el Excel (default AP).")
    ap.add_argument("--bank", required=True, help="Excel del banco BBVA (input)")
    ap.add_argument("--bank-sheet", default="Historico", help="Hoja del banco (default Historico)")
    ap.add_argument("--bank-header", type=int, default=14, help="Fila de cabecera 1-based en banco (default 14)")
    ap.add_argument("--filter-by", choices=["entry","exit"], default="entry", help="Filtrar por 'entry' (entrada) o 'exit' (salida) para mes/año")
    ap.add_argument("--month", type=int, required=True, help="Mes a filtrar (1-12)")
    ap.add_argument("--year", type=int, required=True, help="Año a filtrar (e.g., 2025)")
    ap.add_argument("--date-window-days", type=int, default=-1, help="Ventana en días para fecha bancaria; <0 = sin límite (default)")
    ap.add_argument("--out-liq", default="Liquidaciones_por_caso.xlsx", help="Ruta de salida para liquidaciones por caso")
    ap.add_argument("--out-recon", default="Conciliacion_Banco_vs_Liquidaciones.xlsx", help="Ruta de salida para conciliación")
    args = ap.parse_args()

    # Leer liquidaciones
    liq = pd.read_excel(args.liq, header=max(args.liq_header-1,0))
    liq = ensure_unique_columns(liq)

    # Forzar Portal desde columna indicada si existe
    if args.portal_col in liq.columns:
        liq["Portal"] = liq[args.portal_col].astype(str)

    liq = normalize_columns(liq)

    # Fechas
    liq["Fecha entrada"] = pd.to_datetime(liq.get("Fecha entrada"), errors="coerce", dayfirst=True)
    liq["Fecha salida"]  = pd.to_datetime(liq.get("Fecha salida"), errors="coerce", dayfirst=True)

    # Filtro por mes/año
    if args.filter_by == "entry":
        mask = (liq["Fecha entrada"].dt.month == args.month) & (liq["Fecha entrada"].dt.year == args.year)
    else:
        mask = (liq["Fecha salida"].dt.month == args.month) & (liq["Fecha salida"].dt.year == args.year)
    liq_sel = liq[mask].copy()

    # Procesar por casos (con la regla de IVA Booking según caso)
    out = process_generic(liq_sel)

    # Exportar liquidaciones por caso
    export_liquidaciones_by_case(out, Path(args.out_liq), (CASE1_AMENITIES, CASE2_AMENITIES, CASE3_CLEANING_AMEN, CASE4_PROPS, CASE5_AMENITIES))

    # Leer banco
    bank = pd.read_excel(args.bank, sheet_name=args.bank_sheet, header=max(args.bank_header-1,0))
    bank = ensure_unique_columns(bank)
    col_fecha = "F. CONTABLE" if "F. CONTABLE" in bank.columns else bank.columns[0]
    col_conc = "CONCEPTO" if "CONCEPTO" in bank.columns else bank.columns[1]
    col_bene = "BENEFICIARIO/ORDENANTE" if "BENEFICIARIO/ORDENANTE" in bank.columns else bank.columns[2]
    col_imp  = "IMPORTE" if "IMPORTE" in bank.columns else bank.columns[3]
    bank_df = pd.DataFrame({
        "Fecha": pd.to_datetime(bank[col_fecha], errors="coerce", dayfirst=True),
        "Concepto": bank[col_conc].astype(str).fillna("").str.strip(),
        "Beneficiario/Ordenante": bank[col_bene].astype(str).fillna("").str.strip(),
        "Importe": pd.to_numeric(bank[col_imp], errors="coerce")
    }).dropna(subset=["Fecha","Importe"])

    # Conciliar por Pago recibido
    window = args.date_window_days if args.date_window_days is not None else -1
    df_match, df_bank_un, df_liq_un = reconcile_by_pago_recibido(out, bank_df, date_window_days=window)

    # Exportar conciliación
    export_reconciliation(df_match, df_bank_un, df_liq_un, Path(args.out_recon))

    print("OK")
    print(f"Liquidaciones: {args.out_liq}")
    print(f"Conciliación: {args.out_recon}")

if __name__ == "__main__":
    main()
