import streamlit as st
import pandas as pd
import numpy as np
from datetime import date, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import re

st.set_page_config(page_title="LIQUIDACIONES (Casos 1‚Äì5) + Conciliaci√≥n bancaria avanzada [v9 estricto]", page_icon="üè¶", layout="wide")

# ========= Utilidades de formato =========
MONEY_COLS_CANON = {
    "Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisi√≥n portal",
    "Honorarios Florit","Gasto limpieza","Amenities","Total Gastos","Pago al propietario",
    "Pago recibido","IVA del alquiler"
}
NIGHTS_COL = "Noches ocupadas"

def ensure_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    counts, new_cols = {}, []
    for c in df.columns:
        name = str(c)
        n = counts.get(name, 0)
        new_cols.append(name if n == 0 else f"{name}.{n}")
        counts[name] = n + 1
    out = df.copy()
    out.columns = new_cols
    return out

def _first_existing(df, candidates):
    norm_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        k = str(cand).strip().lower()
        if k in norm_map:
            return norm_map[k]
    return None

def ensure_required(df, required, ctx=""):
    missing = [c for c in required if c not in df.columns]
    if missing:
        st.error(f"Faltan columnas requeridas: {missing} en {ctx}. Ajusta el archivo o usa el modo por letras.")
        st.stop()

def base_name(colname: str) -> str:
    return re.sub(r"\.\d+$", "", str(colname)).strip()

def is_money_col(colname: str) -> bool:
    return base_name(colname) in MONEY_COLS_CANON

def is_nights_col(colname: str) -> bool:
    return base_name(colname).lower() == NIGHTS_COL.lower()

def fmt_number_for_ui(colname: str, x):
    if is_nights_col(colname):
        try:
            return f"{int(round(float(x)))}"
        except Exception:
            return x
    if is_money_col(colname):
        try:
            s = f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            return f"{s} ‚Ç¨"
        except Exception:
            return x
    try:
        return f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return x

def find_col(df: pd.DataFrame, base: str):
    for c in df.columns:
        if base_name(c).lower() == base.strip().lower():
            return c
    return None

def show_table_es_grouped(df: pd.DataFrame, title: str, group_col: str = "Alojamiento"):
    st.subheader(title)

    if group_col not in df.columns:
        view = df.copy()
        total = {c: (view[c].sum() if pd.api.types.is_numeric_dtype(view[c]) else "") for c in view.columns}
        view = pd.concat([view, pd.DataFrame([total], index=["TOTAL"])], axis=0)

        view_fmt = view.copy()
        for c in view_fmt.columns:
            if pd.api.types.is_numeric_dtype(view[c]) or is_money_col(c) or is_nights_col(c):
                view_fmt[c] = view_fmt[c].apply(lambda v: fmt_number_for_ui(c, v))

        def highlight_total(row):
            return ["font-weight: bold;" if row.name == "TOTAL" else "" for _ in row]

        st.dataframe(view_fmt.style.apply(highlight_total, axis=1), use_container_width=True)
        return

    for aloj, subdf in df.groupby(group_col):
        st.markdown(f"**{aloj}**")
        block = subdf.copy()
        total = {c: (block[c].sum() if pd.api.types.is_numeric_dtype(block[c]) else "") for c in block.columns}
        block = pd.concat([block, pd.DataFrame([total], index=["TOTAL"])], axis=0)

        block_fmt = block.copy()
        for c in block_fmt.columns:
            if pd.api.types.is_numeric_dtype(block[c]) or is_money_col(c) or is_nights_col(c):
                block_fmt[c] = block_fmt[c].apply(lambda v: fmt_number_for_ui(c, v))

        def highlight_total(row):
            return ["font-weight: bold;" if row.name == "TOTAL" else "" for _ in row]

        st.dataframe(block_fmt.style.apply(highlight_total, axis=1), use_container_width=True)
        st.divider()

# ========= Normalizaci√≥n =========
LETTER_MAP_DEFAULT = {
    "W": "Alojamiento",
    "D": "Fecha entrada",
    "F": "Fecha salida",
    "H": "Noches ocupadas",
    "I": "Ingreso alojamiento",
    "L": "Ingreso limpieza",    # mapeo fuerte: tarifa limpieza en L
    "O": "Total ingresos",
    "AP": "Portal",
    "AR": "Comisi√≥n portal",
    "AL": "IVA del alquiler",
}

def letters_to_idx(letter):
    s = letter.upper()
    n = 0
    for ch in s:
        if not ('A' <= ch <= 'Z'): return None
        n = n*26 + (ord(ch)-ord('A')+1)
    return n-1

def normalize_columns_by_letters(df, letter_map=LETTER_MAP_DEFAULT):
    out = df.copy()
    cols = list(out.columns)
    rename = {}
    for L, std in letter_map.items():
        i = letters_to_idx(L)
        if i is not None and i < len(cols):
            rename[cols[i]] = std
    out.rename(columns=rename, inplace=True)
    return normalize_columns(out)

def normalize_columns(df):
    out = df.copy()
    col_aloj = _first_existing(out, ["Nombre alojamiento","Alojamiento","Nombre del alojamiento","Nombre Alojamiento"])
    col_fent = _first_existing(out, ["Fecha entrada","Fecha de entrada"])
    col_fsal = _first_existing(out, ["Fecha salida","Fecha de salida"])
    col_noch = _first_existing(out, ["Noches","noches","Noches ocupadas"])
    col_alq  = _first_existing(out, ["Alquiler con tasas","Ingreso alojamiento","Importe alojamiento"])
    col_ext  = _first_existing(out, [
        "Ingreso limpieza","Tarifa limpieza","Limpieza","Importe limpieza",
        "Extras con tasas","Gastos de limpieza","Gasto limpieza"
    ])
    col_tot  = _first_existing(out, ["Total reserva con tasas","Total ingresos","Total"])
    col_port = _first_existing(out, ["Web origen","Portal","Canal","Fuente"])
    col_comi = _first_existing(out, ["Comisi√≥n Portal/Intermediario: Comisi√≥n calculada","Comisi√≥n portal","Comisi√≥n"])
    col_ivaal= _first_existing(out, ["IVA del alojamiento","IVA alojamiento","IVA del alquiler"])

    rename = {}
    if col_aloj: rename[col_aloj] = "Alojamiento"
    if col_fent: rename[col_fent] = "Fecha entrada"
    if col_fsal: rename[col_fsal] = "Fecha salida"
    if col_noch: rename[col_noch] = "Noches ocupadas"
    if col_alq:  rename[col_alq]  = "Ingreso alojamiento"
    if col_ext:  rename[col_ext]  = "Ingreso limpieza"
    if col_tot:  rename[col_tot]  = "Total ingresos"
    if col_port: rename[col_port] = "Portal"
    if col_comi: rename[col_comi] = "Comisi√≥n portal"
    if col_ivaal:rename[col_ivaal]= "IVA del alquiler"

    out.rename(columns=rename, inplace=True)

    # Tipado
    for c in ["Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisi√≥n portal","IVA del alquiler","Noches ocupadas"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)

    for c in ["Fecha entrada","Fecha salida"]:
        if c in out.columns:
            out[c] = pd.to_datetime(out[c], errors="coerce", dayfirst=True)

    if "Alojamiento" in out.columns:
        out["Alojamiento"] = out["Alojamiento"].astype(str).str.strip().str.upper()
    if "Noches ocupadas" in out.columns:
        out["Noches ocupadas"] = pd.to_numeric(out["Noches ocupadas"], errors="coerce").fillna(0).round(0).astype(int)

    return out

# ========= Reglas de casos =========
case1_percent_amenities = {
    "APOLO 180": (0.20, 12.04), "ALMIRANTE 01": (0.22, 11.33), "ALMIRANTE 02": (0.22, 11.33),
    "CADIZ": (0.20, 9.11), "DENIA 61": (0.20, 10.96), "DOLORES ALCAYDE 04": (0.20, 11.33),
    "DR.LLUCH": (0.20, 11.16), "ERUDITO": (0.20, 13.37), "GOZALBO": (0.20, 15.25),
    "LA ELIANA": (0.20, 15.25), "MORAIRA": (0.25, 11.33), "NAPOLES Y SICILIA": (0.25, 0.00),
    "OLIVERETA 5": (0.20, 0.00), "OVE 01": (0.18, 0.00), "OVE 02": (0.18, 0.00),
    "QUART I": (0.20, 9.09), "QUART II": (0.20, 9.09), "SAN LUIS": (0.20, 11.02),
    "SERRANOS": (0.20, 13.37), "SEVILLA": (0.18, 9.45), "TUNDIDORES": (0.20, 7.85),
    "VALLE": (0.20, 11.33),
}
case1_props = set(case1_percent_amenities.keys())

case2_percent_amenities = {
    "VISITACION": (0.20, 14.88),
    "PADRE PORTA 6": (0.20, 12.09), "PADRE PORTA 7": (0.20, 12.09), "PADRE PORTA 8": (0.20, 12.09),
    "PADRE PORTA 9": (0.20, 12.09), "PADRE PORTA 10": (0.20, 12.09),
    "LLADRO Y MALLI 00": (0.20, 9.45), "LLADRO Y MALLI 01": (0.20, 9.45), "LLADRO Y MALLI 02": (0.20, 9.45),
    "LLADRO Y MALLI 03": (0.20, 9.45), "LLADRO Y MALLI 04": (0.20, 9.45),
    "APOLO 29": (0.20, 11.58), "APOLO 197": (0.20, 17.40),
}
case2_props = set(case2_percent_amenities.keys())

case3_cleaning_amenities = {
    "ZAPATEROS 10-2": (0.20, 60.00, 15.24),
    "ZAPATEROS 10-6": (0.20, 75.00, 15.24),
    "ZAPATEROS 10-8": (0.20, 75.00, 15.24),
    "ZAPATEROS 12-5": (0.20, 60.00, 11.33),
    "ALFARO": (0.20, 80.00, 14.88),
}
case3_props = set(case3_cleaning_amenities.keys())

case4_props = {
    "SERRERIA 04", "SERRERIA 05", "RETOR A", "RETOR B",
    "PASAJE ANGELES Y FEDERICO 01", "PASAJE ANGELES Y FEDERICO 02", "PASAJE ANGELES Y FEDERICO 03",
    "MALILLA 05", "MALILLA 06", "MALILLA 07", "MALILLA 08", "MALILLA 14", "MALILLA 15",
    "BENICALAP 01", "BENICALAP 02", "BENICALAP 03", "BENICALAP 04", "BENICALAP 05", "BENICALAP 06"
}

case5_percent_amenities = {
    "HOMERO 01": (0.20, 0.00), "HOMERO 02": (0.20, 0.00),
    "CARCAIXENT 01": (0.20, 8.60), "CARCAIXENT 02": (0.20, 8.60)
}
case5_props = set(case5_percent_amenities.keys())

APOLO_ONLY = {"APOLO 29", "APOLO 197"}

def props_for_case(case):
    if case == 1: return case1_props
    if case == 2: return case2_props
    if case == 3: return case3_props
    if case == 4: return case4_props
    if case == 5: return case5_props
    return set()

# ========= Reglas transversales =========
def apply_commission_vat_by_scope(df: pd.DataFrame, vat_pct: float, treat_empty_as_booking: bool, skip_booking_vat: bool,
                                  scope_mask: pd.Series | None = None) -> tuple[pd.DataFrame, int]:
    out = df.copy()
    portal_col = "Portal"; commission_col = "Comisi√≥n portal"
    if portal_col not in out.columns or commission_col not in out.columns:
        return out, 0

    ser_portal = out[portal_col]
    if isinstance(ser_portal, pd.DataFrame):
        ser_portal = ser_portal.iloc[:, 0]
    ser_portal = ser_portal.astype("string").fillna("")

    out[commission_col] = pd.to_numeric(out[commission_col], errors="coerce").fillna(0.0)
    mask_booking = ser_portal.str.lower().str.contains("booking", na=False)
    mask_empty   = ser_portal.str.strip().eq("")
    warn_count = int(((mask_empty) & (out[commission_col] > 0)).sum())

    if skip_booking_vat or vat_pct == 0:
        return out, warn_count

    mult = 1 + (float(vat_pct) / 100.0)
    mask_scope = scope_mask if scope_mask is not None else pd.Series(True, index=out.index)
    out.loc[(mask_booking & mask_scope), commission_col] *= mult
    if treat_empty_as_booking:
        out.loc[(mask_empty & mask_scope), commission_col] *= mult

    return out, warn_count

# ========= Procesadores =========
def process_case1(df, treat_empty_as_booking=False, skip_booking_vat=False, vat_pct=21.0):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisi√≥n portal","Portal"], "Caso 1")
    scope = pd.Series(True, index=df.index)
    df, warn_count = apply_commission_vat_by_scope(df, vat_pct, treat_empty_as_booking, skip_booking_vat, scope)

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = case1_percent_amenities.get(key,(0.20,0.0))[0]
        return float(r.get("Ingreso alojamiento",0.0)) * pct * 1.21

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case1_percent_amenities.get(key,(0.20,0.0))[1])

    out = df.copy()
    out["Honorarios Florit"] = out.apply(honorarios, axis=1).round(2)
    out["Gasto limpieza"]   = pd.to_numeric(out.get("Ingreso limpieza", 0.0), errors="coerce").fillna(0.0).round(2)
    out["Amenities"]        = out.apply(amenities, axis=1).round(2)
    out["Total Gastos"]     = (out[["Comisi√≥n portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)).round(2)
    out["Pago al propietario"] = (out["Total ingresos"] - out["Total Gastos"]).round(2)
    out["Pago recibido"]    = (out["Total ingresos"] - out["Comisi√≥n portal"]).round(2)

    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","Ingreso limpieza",
            "Total ingresos","Portal","Comisi√≥n portal","Honorarios Florit","Gasto limpieza","Amenities",
            "Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

def process_case2(df, treat_empty_as_booking=False, skip_booking_vat=False, vat_pct=21.0, only_apolo=True):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Total ingresos","Comisi√≥n portal","Portal"], "Caso 2")
    mask_apolo = df["Alojamiento"].astype(str).str.upper().isin(APOLO_ONLY) if only_apolo else pd.Series(True, index=df.index)
    df, warn_count = apply_commission_vat_by_scope(df, vat_pct, treat_empty_as_booking, skip_booking_vat, scope_mask=mask_apolo)

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = case2_percent_amenities.get(key,(0.20,0.0))[0]
        ingreso = float(r.get("Ingreso alojamiento",0.0))
        iva = ingreso - (ingreso / 1.10)
        base = ingreso - iva
        return base * pct * 1.21

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case2_percent_amenities.get(key,(0.20,0.0))[1])

    out = df.copy()
    out["IVA del alquiler"] = pd.to_numeric(out["Ingreso alojamiento"], errors="coerce").fillna(0.0) - (pd.to_numeric(out["Ingreso alojamiento"], errors="coerce").fillna(0.0) / 1.10)
    out["Honorarios Florit"] = out.apply(honorarios, axis=1).round(2)
    out["Gasto limpieza"]   = pd.to_numeric(out.get("Ingreso limpieza", 0.0), errors="coerce").fillna(0.0).round(2)
    out["Amenities"]        = out.apply(amenities, axis=1).round(2)
    out["Total Gastos"]     = (out[["Comisi√≥n portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)).round(2)
    out["Pago al propietario"] = (out["Total ingresos"] - out["Total Gastos"]).round(2)
    out["Pago recibido"]    = (out["Total ingresos"] - out["Comisi√≥n portal"]).round(2)

    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","IVA del alquiler",
            "Ingreso limpieza","Total ingresos","Portal","Comisi√≥n portal","Honorarios Florit","Gasto limpieza",
            "Amenities","Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

def process_case3(df, treat_empty_as_booking=False, skip_booking_vat=False, vat_pct=21.0):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Total ingresos","Comisi√≥n portal","Portal"], "Caso 3")
    scope = pd.Series(True, index=df.index)
    df, warn_count = apply_commission_vat_by_scope(df, vat_pct, treat_empty_as_booking, skip_booking_vat, scope)

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = case3_cleaning_amenities.get(key,(0.20,0.0,0.0))[0]
        base = float(r.get("Ingreso alojamiento",0.0)) - float(r.get("Comisi√≥n portal",0.0))
        return base * pct * 1.21

    def gasto_limpieza(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case3_cleaning_amenities.get(key,(0.20,0.0,0.0))[1])

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case3_cleaning_amenities.get(key,(0.20,0.0,0.0))[2])

    out = df.copy()
    out["Honorarios Florit"] = out.apply(honorarios, axis=1).round(2)
    out["Gasto limpieza"]   = out.apply(gasto_limpieza, axis=1).round(2)
    out["Amenities"]        = out.apply(amenities, axis=1).round(2)
    out["Total Gastos"]     = (out[["Comisi√≥n portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)).round(2)
    out["Pago al propietario"] = (out["Total ingresos"] - out["Total Gastos"]).round(2)
    out["Pago recibido"]    = (out["Total ingresos"] - out["Comisi√≥n portal"]).round(2)

    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","Ingreso limpieza",
            "Total ingresos","Portal","Comisi√≥n portal","Honorarios Florit","Gasto limpieza","Amenities",
            "Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

def process_case4(df, treat_empty_as_booking=False, skip_booking_vat=False, vat_pct=0.0):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Comisi√≥n portal"], "Caso 4")
    df["Portal"] = df.get("Portal", pd.Series([""]*len(df)))
    scope = pd.Series(True, index=df.index)
    df, warn_count = apply_commission_vat_by_scope(df, vat_pct, treat_empty_as_booking, skip_booking_vat, scope)

    out = df.copy()
    ingreso = pd.to_numeric(out.get("Ingreso alojamiento", 0.0), errors="coerce").fillna(0.0)
    out["IVA del alquiler"] = ingreso - (ingreso / 1.10)

    def honorarios(r):
        base = float(r.get("Ingreso alojamiento",0.0)) - float(r.get("IVA del alquiler",0.0)) - float(r.get("Comisi√≥n portal",0.0))
        return base * 0.20

    out["Honorarios Florit"] = out.apply(honorarios, axis=1).round(2)

    out["Pago al propietario"] = (
        pd.to_numeric(out.get("Ingreso alojamiento",0.0), errors="coerce").fillna(0.0)
        - pd.to_numeric(out.get("IVA del alquiler",0.0), errors="coerce").fillna(0.0)
        - pd.to_numeric(out.get("Comisi√≥n portal",0.0), errors="coerce").fillna(0.0)
        - pd.to_numeric(out.get("Honorarios Florit",0.0), errors="coerce").fillna(0.0)
    ).round(2)

    out["Pago recibido"] = (
        pd.to_numeric(out.get("Ingreso alojamiento",0.0), errors="coerce").fillna(0.0)
        + pd.to_numeric(out.get("Ingreso limpieza",0.0), errors="coerce").fillna(0.0)
        - pd.to_numeric(out.get("Comisi√≥n portal",0.0), errors="coerce").fillna(0.0)
    ).round(2)

    cols = [
        "Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas",
        "Ingreso alojamiento","IVA del alquiler","Ingreso limpieza","Total ingresos",
        "Portal","Comisi√≥n portal","Honorarios Florit","Pago al propietario","Pago recibido"
    ]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

def process_case5(df, treat_empty_as_booking=False, skip_booking_vat=False, vat_pct=0.0):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Total ingresos","Comisi√≥n portal"], "Caso 5")
    df["Portal"] = df.get("Portal", pd.Series([""]*len(df)))
    scope = pd.Series(True, index=df.index)
    df, warn_count = apply_commission_vat_by_scope(df, vat_pct, treat_empty_as_booking, skip_booking_vat, scope)

    out = df.copy()
    ingreso = pd.to_numeric(out.get("Ingreso alojamiento", 0.0), errors="coerce").fillna(0.0)
    out["IVA del alquiler"] = ingreso - (ingreso / 1.10)

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = case5_percent_amenities.get(key,(0.20,0.0))[0]
        base = float(r.get("Ingreso alojamiento",0.0)) - float(r.get("IVA del alquiler",0.0)) - float(r.get("Comisi√≥n portal",0.0))
        return base * pct * 1.21

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(case5_percent_amenities.get(key,(0.20,0.0))[1])

    out["Honorarios Florit"] = out.apply(honorarios, axis=1).round(2)
    out["Gasto limpieza"]   = pd.to_numeric(out.get("Ingreso limpieza", 0.0), errors="coerce").fillna(0.0).round(2)
    out["Amenities"]        = out.apply(amenities, axis=1).round(2)
    out["Total Gastos"]     = (out[["Comisi√≥n portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)).round(2)
    out["Pago al propietario"] = (out["Total ingresos"] - out["Total Gastos"]).round(2)
    out["Pago recibido"]    = (out["Total ingresos"] - out["Comisi√≥n portal"]).round(2)

    cols = [
        "Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas",
        "Ingreso alojamiento","IVA del alquiler","Ingreso limpieza","Total ingresos","Portal","Comisi√≥n portal",
        "Honorarios Florit","Gasto limpieza","Amenities","Total Gastos","Pago al propietario","Pago recibido"
    ]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

processors = {1: process_case1, 2: process_case2, 3: process_case3, 4: process_case4, 5: process_case5}

# ========= Exportaci√≥n Excel =========
BORDER_THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

def write_grouped_sheet(ws, df):
    cols = list(df.columns)

    def write_table(start_row, subdf):
        # Cabecera
        for j, col in enumerate(cols, start=1):
            cell = ws.cell(row=start_row, column=j, value=col)
            cell.font = Font(bold=True)
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Filas
        for i, (_, row) in enumerate(subdf.iterrows(), start=1):
            for j, col in enumerate(cols, start=1):
                val = row[col]
                c = ws.cell(row=start_row+i, column=j, value=val)
                c.border = BORDER_THIN
                if isinstance(val, (int, float)) and not pd.isna(val):
                    if is_nights_col(col):
                        c.number_format = "0"
                    elif is_money_col(col):
                        c.number_format = '#.##0,00" ‚Ç¨"'
                    else:
                        c.number_format = "#.##0,00"
                else:
                    c.alignment = Alignment(wrap_text=True)
        # Sumatorios en negrita
        sum_row = start_row + len(subdf) + 1
        ws.cell(row=sum_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=sum_row, column=1).border = BORDER_THIN
        for j, col in enumerate(cols, start=1):
            if j == 1:
                continue
            if pd.api.types.is_numeric_dtype(subdf[col]):
                top = start_row+1
                bottom = start_row+len(subdf)
                formula = f"=SUM({get_column_letter(j)}{top}:{get_column_letter(j)}{bottom})"
                c = ws.cell(row=sum_row, column=j, value=formula)
                c.font = Font(bold=True)
                c.border = BORDER_THIN
                if is_nights_col(col):
                    c.number_format = "0"
                elif is_money_col(col):
                    c.number_format = '#.##0,00" ‚Ç¨"'
                else:
                    c.number_format = "#.##0,00"
            else:
                ws.cell(row=sum_row, column=j, value="").border = BORDER_THIN
        return sum_row + 2

    current_row = 1
    if "Alojamiento" in df.columns:
        for aloj, subdf in df.groupby("Alojamiento"):
            ws.cell(row=current_row, column=1, value=str(aloj)).font = Font(bold=True, size=12)
            current_row += 1
            current_row = write_table(current_row, subdf)
    else:
        current_row = write_table(current_row, df)

    # Auto-ancho
    for j, col in enumerate(cols, start=1):
        max_len = len(str(col))
        for r in range(1, ws.max_row+1):
            v = ws.cell(row=r, column=j).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(j)].width = min(max_len+2, 45)

def build_excel_single(df_final, filename="Liquidacion.xlsx"):
    wb = Workbook(); ws = wb.active; ws.title = "Liquidaci√≥n"
    write_grouped_sheet(ws, df_final)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    st.download_button("üì• Descargar Excel (Liquidaci√≥n)", bio.getvalue(),
                       file_name=filename,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def build_excel_multi(dfs_by_case: dict, filename: str):
    wb = Workbook(); first = True
    for case_label, df_final in dfs_by_case.items():
        if first:
            ws = wb.active; ws.title = case_label; first = False
        else:
            ws = wb.create_sheet(title=case_label)
        write_grouped_sheet(ws, df_final)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    st.download_button("üì• Descargar Excel (Todos los casos)", bio.getvalue(),
                       file_name=filename,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ========= UI: LIQUIDACIONES =========
st.title("üè¶ Conciliaci√≥n bancaria + üìä LIQUIDACIONES (Casos 1‚Äì5) [v9 estricto]")
st.caption("Primero genera las liquidaciones del per√≠odo. Luego sube el extracto bancario y concilia.")

with st.sidebar:
    st.header("Par√°metros de liquidaci√≥n")
    c1, c2 = st.columns(2)
    with c1:
        start_date = st.date_input("Desde", value=date(date.today().year, date.today().month, 1))
    with c2:
        end_date   = st.date_input("Hasta",  value=date(date.today().year, date.today().month, 28))
    st.divider()
    case_choice = st.radio("Caso", ["Todos", 1,2,3,4,5], horizontal=False)
    st.checkbox("Lectura por letras (fallback)", value=False, key="by_letters")
    st.caption("Mapeo: W, D, F, H, I, J/L (L limpia), O, AP, AR, AL.")
    st.divider()
    st.subheader("IVA comisi√≥n por caso (Booking)")
    col_v1, col_v2 = st.columns(2)
    with col_v1:
        vat_case1 = st.number_input("Caso 1 (%)", min_value=0.0, max_value=30.0, value=21.0, step=0.5)
        vat_case3 = st.number_input("Caso 3 (%)", min_value=0.0, max_value=30.0, value=21.0, step=0.5)
        vat_case5 = st.number_input("Caso 5 (%)", min_value=0.0, max_value=30.0, value=0.0, step=0.5)
    with col_v2:
        vat_case2 = st.number_input("Caso 2 (%)", min_value=0.0, max_value=30.0, value=21.0, step=0.5)
        vat_case4 = st.number_input("Caso 4 (%)", min_value=0.0, max_value=30.0, value=0.0, step=0.5)
        only_apolo_c2 = st.checkbox("Caso 2: aplicar solo a APOLO 29/197", value=True)
    treat_empty_as_booking = st.checkbox("Tratar portal vac√≠o como Booking (aplicar IVA comisi√≥n)", value=False)
    skip_booking_vat = st.checkbox("No a√±adir IVA a comisi√≥n de Booking (ya viene con IVA)", value=False)
    generate = st.button("Generar liquidaci√≥n")

file = st.file_uploader("Sube el archivo de reservas (.xlsx)", type=["xlsx"], key="reservas_upl")

# ========= Generaci√≥n Liquidaciones =========
def normalize_liq_for_period(df_norm, start_date, end_date):
    if "Fecha entrada" in df_norm.columns:
        mask = (df_norm["Fecha entrada"] >= pd.to_datetime(start_date)) & (df_norm["Fecha entrada"] <= pd.to_datetime(end_date))
        df_norm = df_norm[mask]
    return df_norm

if generate:
    if not file:
        st.error("Sube primero el archivo de reservas (.xlsx).")
        st.stop()

    df_in = pd.read_excel(file, header=0)
    df_in = ensure_unique_columns(df_in)
    df_norm = normalize_columns_by_letters(df_in) if st.session_state.by_letters else normalize_columns(df_in)
    df_norm = ensure_unique_columns(df_norm)
    df_norm = normalize_liq_for_period(df_norm, start_date, end_date)

    if "Ingreso limpieza" in df_norm.columns:
        limp = pd.to_numeric(df_norm["Ingreso limpieza"], errors="coerce").fillna(0)
        if (limp > 300).any():
            st.warning("Detectadas tarifas de limpieza > 300 ‚Ç¨. Verifica que la columna L est√© mapeada como 'Ingreso limpieza' o activa el modo por letras.")

    def run_case(case_no):
        df_case = df_norm.copy()
        props = props_for_case(case_no)
        if props and "Alojamiento" in df_case.columns:
            df_case = df_case[df_case["Alojamiento"].isin(props)]
        vat_map = {1: vat_case1, 2: vat_case2, 3: vat_case3, 4: vat_case4, 5: vat_case5}
        if case_no == 2:
            out, warn = processors[case_no](df_case, treat_empty_as_booking=treat_empty_as_booking, skip_booking_vat=skip_booking_vat, vat_pct=vat_map[case_no], only_apolo=only_apolo_c2)
        elif case_no in (1,3,4,5):
            out, warn = processors[case_no](df_case, treat_empty_as_booking=treat_empty_as_booking, skip_booking_vat=skip_booking_vat, vat_pct=vat_map[case_no])
        else:
            out, warn = processors[case_no](df_case)
        if NIGHTS_COL in out.columns:
            out[NIGHTS_COL] = pd.to_numeric(out[NIGHTS_COL], errors="coerce").fillna(0).round(0).astype(int)
        for c in out.columns:
            if c != NIGHTS_COL and pd.api.types.is_numeric_dtype(out[c]):
                out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0).round(2)
        return out, warn

    if case_choice == "Todos":
        dfs = {}; total_warns = 0
        for c in [1,2,3,4,5]:
            df_out, warn = run_case(c)
            total_warns += warn
            df_out = df_out.sort_values(by=[col for col in ["Alojamiento","Fecha entrada"] if col in df_out.columns])
            dfs[f"Caso {c}"] = df_out
        st.success(f"Liquidaci√≥n generada (Todos) ‚Ä¢ {start_date.strftime('%d/%m/%Y')}‚Äì{end_date.strftime('%d/%m/%Y')}")

        for label, df_show in dfs.items():
            show_table_es_grouped(df_show, f"{label} ‚Äî Tabla de liquidaciones")

        file_name = f"Liquidaciones_TODOS_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        build_excel_multi(dfs, filename=file_name)

        st.session_state["df_liq_all"] = pd.concat(dfs.values(), ignore_index=True, sort=False)
        st.session_state["df_liq_label"] = "Todos"
        if total_warns > 0 and not treat_empty_as_booking:
            st.warning(f"Hay {total_warns} reservas con comisi√≥n > 0 pero portal vac√≠o en alguno de los casos. Si deben ser Booking, marca la opci√≥n correspondiente y vuelve a generar.")
    else:
        case_no = int(case_choice)
        df_out, warn = run_case(case_no)
        df_out = df_out.sort_values(by=[col for col in ["Alojamiento","Fecha entrada"] if col in df_out.columns])

        st.success(f"Liquidaci√≥n generada (Caso {case_no}) ‚Ä¢ {start_date.strftime('%d/%m/%Y')}‚Äì{end_date.strftime('%d/%m/%Y')}")
        show_table_es_grouped(df_out, "Tabla de liquidaciones")

        aloj_col = find_col(df_out, "Alojamiento")
        pago_col = find_col(df_out, "Pago al propietario")
        if aloj_col is not None and pago_col is not None:
            pagos = (df_out[[aloj_col, pago_col]].groupby(aloj_col, as_index=False)[pago_col]
                     .sum().round(2).sort_values(aloj_col))
            pagos.rename(columns={aloj_col: "Alojamiento", pago_col: "Pago al propietario"}, inplace=True)
            pagos_fmt = pagos.copy()
            for c in pagos_fmt.columns:
                if pd.api.types.is_numeric_dtype(pagos_fmt[c]) or is_money_col(c):
                    pagos_fmt[c] = pagos_fmt[c].apply(lambda v: fmt_number_for_ui(c, v))
            st.subheader("üí∏ Pagos por alojamiento (suma)")
            st.dataframe(pagos_fmt, use_container_width=True)

        file_case_name = f"Liquidacion_CASO{case_no}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        build_excel_single(df_out, filename=file_case_name)

        st.session_state["df_liq_all"] = df_out.copy()
        st.session_state["df_liq_label"] = f"Caso {case_no}"
        if warn > 0 and not treat_empty_as_booking:
            st.warning("Hay reservas con comisi√≥n > 0 pero portal vac√≠o. Si deben ser Booking, marca ‚ÄòTratar portal vac√≠o como Booking‚Äô.")

st.divider()
st.header("üîó Conciliaci√≥n bancaria (mejorada)")

# === Subida del banco y opciones de cabecera ===
colA, colB, colC, colD = st.columns([1,1,1,1])
with colA:
    bank_file = st.file_uploader("Sube el extracto bancario (.xlsx)", type=["xlsx"], key="bank_upl")
with colB:
    bank_sheet = st.text_input("Hoja del banco", value="Historico")
with colC:
    header_row_1based = st.number_input("Fila donde empieza el CABECERO", min_value=1, value=14, step=1,
                                        help="Ej.: en el BBVA, los encabezados est√°n en la fila 14.")
with colD:
    strict_cents = st.checkbox("üîí Modo estricto (c√©ntimo exacto)", value=True,
                               help="Usa aritm√©tica en c√©ntimos y tolerancia 0 en importes.")

if bank_file:
    try:
        hdr = int(header_row_1based) - 1
        raw_bank = pd.read_excel(bank_file, sheet_name=bank_sheet, header=hdr)
        raw_bank = ensure_unique_columns(raw_bank)
        st.caption(f"Columnas detectadas: {list(raw_bank.columns)}")
    except Exception as e:
        st.error(f"No se pudo leer el archivo del banco: {e}")
        raw_bank = None
else:
    raw_bank = None

# === Mapeo de columnas del banco ===
if raw_bank is not None:
    st.subheader("Mapeo de columnas del banco")
    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        col_fecha = st.selectbox("Columna de FECHA", options=list(raw_bank.columns), index=(list(raw_bank.columns).index("F. CONTABLE") if "F. CONTABLE" in raw_bank.columns else 0))
    with c2:
        col_conc = st.selectbox("Columna de CONCEPTO", options=list(raw_bank.columns), index=(list(raw_bank.columns).index("CONCEPTO") if "CONCEPTO" in raw_bank.columns else 0))
    with c3:
        col_bene = st.selectbox("Columna BENEF./ORDENANTE", options=list(raw_bank.columns), index=(list(raw_bank.columns).index("BENEFICIARIO/ORDENANTE") if "BENEFICIARIO/ORDENANTE" in raw_bank.columns else 0))
    with c4:
        col_imp  = st.selectbox("Columna de IMPORTE", options=list(raw_bank.columns), index=(list(raw_bank.columns).index("IMPORTE") if "IMPORTE" in raw_bank.columns else 0))
    with c5:
        importe_es_negativo = st.checkbox("Importe viene NEGATIVO en pagos (convertir a positivo)", value=False)

    # Preprocesado del banco
    bank = pd.DataFrame({
        "Fecha": pd.to_datetime(raw_bank[col_fecha], errors="coerce", dayfirst=True),
        "Concepto": raw_bank[col_conc].astype(str).fillna("").str.strip(),
        "Beneficiario/Ordenante": raw_bank[col_bene].astype(str).fillna("").str.strip(),
        "Importe": pd.to_numeric(raw_bank[col_imp], errors="coerce")
    }).dropna(subset=["Fecha","Importe"])

    if importe_es_negativo:
        bank["Importe"] = bank["Importe"].apply(lambda x: -x if x < 0 else x)

    only_neg = st.checkbox("Quedarme solo con pagos (importes negativos)", value=False)
    if only_neg:
        bank = bank[bank["Importe"] < 0].copy()
    st.dataframe(bank.head(20), use_container_width=True)

    st.subheader("Par√°metros de conciliaci√≥n")
    cc1, cc2, cc3 = st.columns(3)
    with cc1:
        if strict_cents:
            st.number_input("Tolerancia en importe (‚Ç¨)", min_value=0.00, value=0.00, step=0.01, format="%.2f", disabled=True, help="Modo estricto: 0 ‚Ç¨")
            amt_tol = 0.00
        else:
            amt_tol = st.number_input("Tolerancia en importe (‚Ç¨)", min_value=0.00, value=0.01, step=0.01, format="%.2f")
    with cc2:
        days_tol = st.number_input("Tolerancia en d√≠as (¬±)", min_value=0, value=2, step=1)
    with cc3:
        liquid_date_col = st.selectbox("Fecha de referencia en liquidaciones", options=["Fecha entrada","Fecha salida"], index=1)

    # Opciones avanzadas de combinaciones
    st.subheader("Opciones avanzadas de combinaciones")
    colx1, colx2, colx3 = st.columns(3)
    with colx1:
        enable_1N = st.checkbox("Habilitar 1‚ÜîN (pago cubre varias reservas)", value=True)
    with colx2:
        enable_N1 = st.checkbox("Habilitar N‚Üî1 (varios pagos cubren una reserva)", value=True)
    with colx3:
        max_k = st.number_input("K m√°ximo por combinaci√≥n", min_value=2, max_value=5, value=3, step=1)
    coly1, coly2 = st.columns(2)
    with coly1:
        restrict_same_portal = True  # Forzado: combinaciones solo dentro del mismo Portal
    with coly2:
        group_by_base_aloj = True  # Forzado: combinaciones solo dentro de la misma base de alojamiento

    booking_same_base = True  # Forzado: no mezclar bases en Booking (y extendido al resto por l√≥gica principal)

    btn_reconcile = st.button("Conciliar pagos ‚Üî liquidaciones (avanzado)")

    # Helpers combinaciones
    def alojamiento_base(s: str) -> str:
        s = (s or "").upper()
        s = re.sub(r"\b\d+\b", "", s)
        s = re.sub(r"\s{2,}", " ", s)
        return s.strip()

    def to_cents(x) -> int:
        try:
            return int(round(float(x) * 100))
        except Exception:
            try:
                return int(round(float(str(x).replace(",", ".")) * 100))
            except Exception:
                return 0

    def subset_sum_indices_cents(values_cents, indices, target_cents, tol_cents, kmax):
        """
        Backtracking en c√©ntimos con poda: busca subconjunto (hasta kmax) cuya suma == target dentro de tol_cents.
        values_cents: lista de importes en c√©ntimos (positivos, ints)
        indices: √≠ndices asociados a cada valor
        """
        best = None
        n = len(values_cents)
        order = sorted(range(n), key=lambda i: abs(values_cents[i]-target_cents))
        vals = [values_cents[i] for i in order]
        idxs = [indices[i] for i in order]

        def backtrack(start, kleft, acc_sum, chosen):
            nonlocal best
            if kleft == 0:
                if abs(acc_sum - target_cents) <= tol_cents:
                    best = chosen[:]
                return
            if best is not None:
                return
            # poda simple por exceso
            if acc_sum - target_cents > tol_cents:
                return
            for p in range(start, len(vals)):
                v = vals[p]
                if acc_sum + v - target_cents > tol_cents and v > 0:
                    continue
                chosen.append(idxs[p])
                backtrack(p+1, kleft-1, acc_sum+v, chosen)
                chosen.pop()
                if best is not None:
                    return

        for size in range(2, kmax+1):
            backtrack(0, size, 0, [])
            if best is not None:
                return best
        return None

    # === Conciliaci√≥n ===
    if btn_reconcile:
        if "df_liq_all" not in st.session_state:
            st.error("Primero genera las liquidaciones arriba (bot√≥n Generar).")
            st.stop()

        liq = st.session_state["df_liq_all"].copy()

        if liquid_date_col not in liq.columns:
            st.error(f"En las liquidaciones no existe la columna '{liquid_date_col}'. Genera de nuevo o elige otra fecha.")
            st.stop()
        ensure_required(liq, ["Pago recibido"], "Conciliaci√≥n")

        # Campos auxiliares
        liq["__FechaRef__"] = pd.to_datetime(liq[liquid_date_col], errors="coerce", dayfirst=True)
        liq["__Pago__"] = pd.to_numeric(liq["Pago recibido"], errors="coerce").round(2)
        liq["__PagoCents__"] = liq["__Pago__"].abs().apply(to_cents)
        liq["__Portal__"] = liq.get("Portal", "").astype(str).str.upper()
        liq["__AlojBase__"] = liq.get("Alojamiento", "").astype(str).apply(alojamiento_base)

        bank2 = bank.copy()
        bank2["__Fecha__"] = pd.to_datetime(bank2["Fecha"], errors="coerce")
        bank2["__ImporteAbs__"] = bank2["Importe"].abs().round(2)
        bank2["__ImporteAbsCents__"] = bank2["__ImporteAbs__"].apply(to_cents)

        used_liq = set()
        used_bank = set()
        matches_rows = []

        tol_cents = 0 if strict_cents else to_cents(amt_tol)

        # 1) 1‚Üî1 b√°sico con tolerancias (en c√©ntimos si estricto)
        for j, rb in bank2.iterrows():
            if j in used_bank: continue
            b_amt_c = int(rb["__ImporteAbsCents__"])
            # candidatos por fecha
            cands = liq.loc[~liq.index.isin(used_liq)].copy()
            cands = cands[
                cands["__FechaRef__"].between(rb["__Fecha__"] - timedelta(days=int(days_tol)),
                                              rb["__Fecha__"] + timedelta(days=int(days_tol)))
            ]
            if cands.empty: 
                continue
            if strict_cents:
                cands["__diff_cents__"] = (cands["__PagoCents__"] - b_amt_c).abs()
                cands = cands[cands["__diff_cents__"] <= tol_cents]
            else:
                b_amt = float(rb["__ImporteAbs__"])
                cands["__diff__"] = (cands["__Pago__"].abs() - b_amt).abs()
                cands = cands[cands["__diff__"] <= float(amt_tol) + 1e-9]
            if cands.empty: 
                continue
            # el mejor (menor Œî importe, luego Œî d√≠as)
            cands["__d_dias__"] = (cands["__FechaRef__"] - rb["__Fecha__"]).abs().dt.days
            sort_cols = ["__diff_cents__","__d_dias__"] if strict_cents else ["__diff__","__d_dias__"]
            cands = cands.sort_values(sort_cols)
            i = cands.index[0]
            used_liq.add(i); used_bank.add(j)
            rf = liq.loc[i]
            matches_rows.append({
                "Tipo": "1‚Üî1",
                "Fecha mov.": rb["__Fecha__"].date() if not pd.isna(rb["__Fecha__"]) else None,
                "Concepto": rb["Concepto"],
                "Benef./Ord.": rb["Beneficiario/Ordenante"],
                "Importe mov.": rb["Importe"],
                "Reservas": str(rf.get("Alojamiento","")),
                "Fecha(s) reserva": str(rf.get("Fecha entrada","")),
                "Portal": rf.get("Portal",""),
                "Pago(s) recibido(s)": round(float(rf.get("Pago recibido", np.nan)),2),
                "Œî d√≠as": int(abs((rf["__FechaRef__"] - rb["__Fecha__"]).days)) if pd.notna(rf["__FechaRef__"]) and pd.notna(rb["__Fecha__"]) else None,
                "Œî importe": round(float(abs(abs(rf["__Pago__"]) - float(rb.get('__ImporteAbs__',0.0)))),2)
            })

        # Helper: preparar candidatos filtrados por pol√≠ticas
        def candidate_liq_for_bank(j, rb):
            cands = liq.loc[~liq.index.isin(used_liq)].copy()
            cands = cands[
                cands["__FechaRef__"].between(rb["__Fecha__"] - timedelta(days=int(days_tol)),
                                              rb["__Fecha__"] + timedelta(days=int(days_tol)))
            ]
            return cands


        # 2) 1‚ÜîN (pago cubre varias reservas) ‚Äî ESTRICTO:
        #    Solo se permiten combinaciones dentro del MISMO Portal y la MISMA base de alojamiento.
        if enable_1N:
            for j, rb in bank2.iterrows():
                if j in used_bank: 
                    continue
                b_amt_c = int(rb["__ImporteAbsCents__"])

                # Candidatos por ventana temporal
                cands = liq.loc[~liq.index.isin(used_liq)].copy()
                cands = cands[
                    cands["__FechaRef__"].between(rb["__Fecha__"] - timedelta(days=int(days_tol)),
                                                  rb["__Fecha__"] + timedelta(days=int(days_tol)))
                ]
                if cands.empty:
                    continue

                # Reducir universo por cercan√≠a temporal
                cands["__d_dias__"] = (cands["__FechaRef__"] - rb["__Fecha__"]).abs().dt.days
                cands = cands.sort_values("__d_dias__").head(120)

                # Agrupar estrictamente por (Portal, Base)
                if "__Portal__" not in cands.columns or "__AlojBase__" not in cands.columns:
                    continue

                solved = False
                for (p, base), sub in cands.groupby(["__Portal__", "__AlojBase__"]):
                    if pd.isna(p) or p == "" or pd.isna(base) or base == "":
                        continue
                    if len(sub) < 2:
                        continue

                    vals_c = sub["__PagoCents__"].abs().astype(int).tolist()
                    idxs   = sub.index.tolist()
                    sol = subset_sum_indices_cents(vals_c, idxs, b_amt_c, tol_cents, int(max_k))
                    if sol:
                        used_bank.add(j)
                        for i in sol: 
                            used_liq.add(i)
                        subsel = liq.loc[sol]
                        reservas = "; ".join(subsel["Alojamiento"].astype(str).tolist())
                        fechas   = "; ".join(subsel["Fecha entrada"].astype(str).tolist()) if "Fecha entrada" in subsel.columns else ""
                        pago_sum = round(float(subsel["Pago recibido"].abs().sum()),2)
                        matches_rows.append({
                            "Tipo": "1‚ÜîN",
                            "Fecha mov.": rb["__Fecha__"].date() if not pd.isna(rb["__Fecha__"]) else None,
                            "Concepto": rb["Concepto"],
                            "Benef./Ord.": rb["Beneficiario/Ordenante"],
                            "Importe mov.": rb["Importe"],
                            "Reservas": reservas,
                            "Fecha(s) reserva": fechas,
                            "Portal": p,
                            "Pago(s) recibido(s)": pago_sum,
                            "Œî d√≠as": int((subsel["__FechaRef__"] - rb["__Fecha__"]).abs().dt.days.max()) if "__FechaRef__" in subsel else None,
                            "Œî importe": round(abs(pago_sum - float(rb["__ImporteAbs__"])),2)
                        })
                        solved = True
                        break  # No buscar m√°s bases para este movimiento
                # Si no hay combinaci√≥n estricta dentro de (portal, base), se deja sin match por 1‚ÜîN y seguir√° al 3) N‚Üî1
        # 3) N‚Üî1 (varios pagos cubren una reserva) con c√©ntimos
        if enable_N1:
            for i, rf in liq.loc[~liq.index.isin(used_liq)].iterrows():
                l_amt_c = int(rf["__PagoCents__"])
                bmask = bank2.index.isin(used_bank) == False
                bsub = bank2.loc[bmask].copy()
                if bsub.empty: 
                    continue
                if pd.notna(rf["__FechaRef__"]):
                    bsub = bsub[
                        bsub["__Fecha__"].between(rf["__FechaRef__"] - timedelta(days=int(days_tol)),
                                                   rf["__FechaRef__"] + timedelta(days=int(days_tol)))
                    ]
                if len(bsub) < 2: 
                    continue
                # Orden preliminar
                bsub["__d_dias__"] = (bsub["__Fecha__"] - rf["__FechaRef__"]).abs().dt.days if pd.notna(rf["__FechaRef__"]) else 0
                bsub = bsub.sort_values(["__d_dias__"]).head(80)
                vals_c = bsub["__ImporteAbsCents__"].abs().astype(int).tolist()
                idxs = bsub.index.tolist()
                sol = subset_sum_indices_cents(vals_c, idxs, l_amt_c, tol_cents, int(max_k))
                if sol:
                    used_liq.add(i)
                    for j in sol: used_bank.add(j)
                    subsel = bank2.loc[sol]
                    reservas = str(rf.get("Alojamiento",""))
                    fechas = str(rf.get("Fecha entrada",""))
                    pago_sum = round(float(subsel["Importe"].abs().sum()),2)
                    matches_rows.append({
                        "Tipo": "N‚Üî1",
                        "Fecha mov.": "; ".join(subsel["__Fecha__"].dt.date.astype(str).tolist()),
                        "Concepto": "; ".join(subsel["Concepto"].astype(str).tolist()),
                        "Benef./Ord.": "; ".join(subsel["Beneficiario/Ordenante"].astype(str).tolist()),
                        "Importe mov.": "; ".join(subsel["Importe"].astype(str).tolist()),
                        "Reservas": reservas,
                        "Fecha(s) reserva": fechas,
                        "Portal": rf.get("Portal",""),
                        "Pago(s) recibido(s)": round(float(rf.get("Pago recibido", np.nan)),2),
                        "Œî d√≠as": int(subsel["__Fecha__"].sub(rf["__FechaRef__"]).abs().dt.days.max()) if pd.notna(rf["__FechaRef__"]) else None,
                        "Œî importe": round(abs(pago_sum - float(rf.get("Pago recibido", 0.0))),2)
                    })

        df_matches = pd.DataFrame(matches_rows)

        # Tablas sin match
        unmatched_bank = bank2.loc[~bank2.index.isin(used_bank), ["Fecha","Concepto","Beneficiario/Ordenante","Importe"]].copy()
        unmatched_liq = liq.loc[~liq.index.isin(used_liq)].copy()

        st.success(f"Conciliaci√≥n realizada ‚Ä¢ Coincidencias: {len(df_matches)}")

        st.subheader("‚úÖ Pagos conciliados")
        if not df_matches.empty:
            st.dataframe(df_matches, use_container_width=True)
        else:
            st.info("No se encontraron coincidencias con los par√°metros actuales.")

        st.subheader("‚ùå Movimientos bancarios sin liquidaci√≥n")
        st.dataframe(unmatched_bank, use_container_width=True)

        st.subheader("‚è≥ Liquidaciones sin pago encontrado")
        cols_show = [c for c in ["Alojamiento","Fecha entrada","Fecha salida","Pago recibido"] if c in unmatched_liq.columns]
        st.dataframe(unmatched_liq[cols_show] if cols_show else unmatched_liq, use_container_width=True)

        # Descargar Excel de conciliaci√≥n
        wb = Workbook()
        ws1 = wb.active; ws1.title = "Conciliados"
        def write_sheet(ws, dfi):
            if dfi is None or len(dfi)==0:
                ws.append(["(sin datos)"]); return
            for j, col in enumerate(dfi.columns, start=1):
                ws.cell(row=1, column=j, value=str(col)).font = Font(bold=True)
            for i, (_, row) in enumerate(dfi.iterrows(), start=2):
                for j, col in enumerate(dfi.columns, start=1):
                    val = row[col]
                    c = ws.cell(row=i, column=j, value=val)
                    if isinstance(val, (int,float)) and not pd.isna(val):
                        if ("Importe" in str(col)) or ("Pago" in str(col)) or ("Œî importe" in str(col)):
                            c.number_format = '#.##0,00" ‚Ç¨"'
                        elif ("Œî d√≠as" in str(col)):
                            c.number_format = '0'
                        else:
                            c.number_format = '#.##0,00'
        write_sheet(ws1, df_matches)
        ws2 = wb.create_sheet("Banco_sin_match"); write_sheet(ws2, unmatched_bank)
        ws3 = wb.create_sheet("Liq_sin_match"); write_sheet(ws3, unmatched_liq[cols_show] if cols_show else unmatched_liq)

        bio = BytesIO(); wb.save(bio); bio.seek(0)
        st.download_button(
            "üì• Descargar Excel de conciliaci√≥n",
            bio.getvalue(),
            file_name=f"Conciliacion_{st.session_state.get('df_liq_label','Casos')}_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
