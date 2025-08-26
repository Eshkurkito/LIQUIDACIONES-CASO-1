
import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

st.set_page_config(page_title="LIQUIDACIONES (Casos 1-5)", page_icon="📊", layout="wide")

# ----------------------
# Utilidades comunes
# ----------------------
def _first_existing(df, candidates):
    norm_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        key = str(cand).strip().lower()
        if key in norm_map:
            return norm_map[key]
    return None

def normalize_columns(df):
    out = df.copy()

    col_aloj = _first_existing(out, ["Alojamiento","Nombre alojamiento","Nombre del alojamiento","Nombre Alojamiento"])
    col_fent = _first_existing(out, ["Fecha entrada","Fecha de entrada"])
    col_fsal = _first_existing(out, ["Fecha salida","Fecha de salida"])
    col_noch = _first_existing(out, ["noches","Noches","Noches ocupadas"])
    col_alq  = _first_existing(out, ["Alquiler con tasas","Ingreso alojamiento","Importe alojamiento"])
    col_ext  = _first_existing(out, ["Extras con tasas","Ingreso limpieza","Limpieza","Importe limpieza"])
    col_tot  = _first_existing(out, ["Total reserva con tasas","Total ingresos","Total"])
    col_port = _first_existing(out, ["Web origen","Portal","Canal","Fuente"])
    col_comi = _first_existing(out, ["Comisión Portal/Intermediario: Comisión calculada","Comisión portal","Comisión"])
    col_ivaal= _first_existing(out, ["IVA del alojamiento","IVA alojamiento"])

    rename_map = {}
    if col_aloj: rename_map[col_aloj] = "Alojamiento"
    if col_fent: rename_map[col_fent] = "Fecha entrada"
    if col_fsal: rename_map[col_fsal] = "Fecha salida"
    if col_noch: rename_map[col_noch] = "Noches ocupadas"
    if col_alq:  rename_map[col_alq]  = "Ingreso alojamiento"
    if col_ext:  rename_map[col_ext]  = "Ingreso limpieza"
    if col_tot:  rename_map[col_tot]  = "Total ingresos"
    if col_port: rename_map[col_port] = "Portal"
    if col_comi: rename_map[col_comi] = "Comisión portal"
    if col_ivaal:rename_map[col_ivaal]= "IVA del alquiler"

    out.rename(columns=rename_map, inplace=True)

    for c in ["Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisión portal","IVA del alquiler","Noches ocupadas"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)

    for c in ["Fecha entrada","Fecha salida"]:
        if c in out.columns:
            out[c] = pd.to_datetime(out[c], errors="coerce", dayfirst=True)

    if "Alojamiento" in out.columns:
        out["Alojamiento"] = out["Alojamiento"].astype(str).str.strip().str.upper()

    return out

def keep_original_dates_as_text(df, date_cols):
    out = df.copy()
    for c in date_cols:
        if c in out.columns:
            out[c] = pd.to_datetime(out[c], errors="coerce", dayfirst=True).dt.strftime("%d/%m/%Y").fillna("")
    return out

def apply_booking_vat_on_commission(df, portal_col="Portal", commission_col="Comisión portal"):
    out = df.copy()
    if portal_col in out.columns and commission_col in out.columns:
        mask = out[portal_col].astype(str).str.lower().str.contains("booking", na=False)
        out.loc[mask, commission_col] = out.loc[mask, commission_col] * 1.21
    return out

def to_excel_grouped(df, group_col="Alojamiento", name="Liquidación"):
    number_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]

    if "Pago recibido" in df.columns and "" not in df.columns:
        idx = list(df.columns).index("Pago recibido")
        df = df.copy()
        df.insert(idx, "", "")

    wb = Workbook()
    ws = wb.active
    ws.title = name

    bold = Font(bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center")

    cols = df.columns.tolist()
    row_cursor = 1

    for group_val, subdf in df.groupby(group_col):
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=row_cursor, column=col_idx, value=col_name)
            cell.font = bold
            cell.border = border
            cell.alignment = center
        row_cursor += 1

        for _, row in subdf.iterrows():
            for col_idx, col_name in enumerate(cols, start=1):
                value = row[col_name]
                cell = ws.cell(row=row_cursor, column=col_idx, value=value)
                cell.border = border
            row_cursor += 1

        for col_idx, col_name in enumerate(cols, start=1):
            if col_name == group_col:
                ws.cell(row=row_cursor, column=col_idx, value="Total").font = bold
            elif col_name in number_cols:
                subtotal = subdf[col_name].sum()
                c = ws.cell(row=row_cursor, column=col_idx, value=float(round(subtotal, 2)))
                c.font = bold
                c.border = border
        row_cursor += 2

    number_format = "#.##0,00"
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value is not None:
                cell.number_format = number_format

    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max(10, min(max_len + 2, 60))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ----------------------
# Configuración por casos
# ----------------------
case1_percent_amenities = {
    "APOLO 180": (0.20, 12.04), "ALMIRANTE 01": (0.22, 11.33), "ALMIRANTE 02": (0.22, 11.33),
    "CADIZ": (0.20, 9.11), "DENIA 61": (0.20, 10.96), "DOLORES ALCAYDE 04": (0.20, 11.33),
    "DR.LLUCH": (0.20, 11.16), "ERUDITO": (0.20, 13.37), "GOZALBO": (0.20, 15.25),
    "LA ELIANA": (0.20, 15.25), "MORAIRA": (0.25, 11.33), "NAPOLES Y SICILIA": (0.25, 0.00),
    "OLIVERETA 5": (0.20, 0.00), "OVE 01": (0.18, 0.00), "OVE 02": (0.18, 0.00),
    "QUART I": (0.20, 9.09), "QUART II": (0.20, 9.09), "SAN LUIS": (0.20, 11.02),
    "SERRANOS": (0.20, 13.37), "SEVILLA": (0.18, 9.45), "TUNDIDORES": (0.20, 7.85),
    "VALLE": (0.20, 11.33),
}
case1_props = set(case1_percent_amenities.keys())

case2_percent_amenities = {
    "VISITACION": (0.20, 14.88),
    "PADRE PORTA 6": (0.20, 12.09), "PADRE PORTA 7": (0.20, 12.09), "PADRE PORTA 8": (0.20, 12.09),
    "PADRE PORTA 9": (0.20, 12.09), "PADRE PORTA 10": (0.20, 12.09),
    "LLADRO Y MALLI 00": (0.20, 9.45), "LLADRO Y MALLI 01": (0.20, 9.45), "LLADRO Y MALLI 02": (0.20, 9.45),
    "LLADRO Y MALLI 03": (0.20, 9.45), "LLADRO Y MALLI 04": (0.20, 9.45),
    "APOLO 29": (0.20, 11.58), "APOLO 197": (0.20, 17.40),
}
case2_props = set(case2_percent_amenities.keys())

case3_cleaning_amenities = {
    "ZAPATEROS 10-2": (0.20, 60.00, 15.24),
    "ZAPATEROS 10-6": (0.20, 75.00, 15.24),
    "ZAPATEROS 10-8": (0.20, 75.00, 15.24),
    "ZAPATEROS 12-5": (0.20, 60.00, 11.33),
    "ALFARO": (0.20, 80.00, 14.88),
}
case3_props = set(case3_cleaning_amenities.keys())

case4_props = {
    "SERRERIA 04", "SERRERIA 05", "RETOR A", "RETOR B",
    "PASAJE ANGELES Y FEDERICO 01", "PASAJE ANGELES Y FEDERICO 02", "PASAJE ANGELES Y FEDERICO 03",
    "MALILLA 05", "MALILLA 06", "MALILLA 07", "MALILLA 08", "MALILLA 14", "MALILLA 15",
    "BENICALAP 01", "BENICALAP 02", "BENICALAP 03", "BENICALAP 04", "BENICALAP 05", "BENICALAP 06"
}
case4_percent_default = 0.20

case5_percent_amenities = {
    "HOMERO 01": (0.20, 0.00), "HOMERO 02": (0.20, 0.00),
    "CARCAIXENT 01": (0.20, 8.60), "CARCAIXENT 02": (0.20, 8.60),
}
case5_props = set(case5_percent_amenities.keys())

# ----------------------
# Procesadores por caso
# ----------------------
def process_case1(df):
    df = normalize_columns(df)
    df = apply_booking_vat_on_commission(df, "Portal", "Comisión portal")
    def honorarios(row):
        pct = case1_percent_amenities.get(row["Alojamiento"], (0.20, 0.0))[0]
        return row["Ingreso alojamiento"] * pct * 1.21
    def amenities(row):
        return case1_percent_amenities.get(row["Alojamiento"], (0.20, 0.0))[1]
    out = df.copy()
    out["Honorarios Florit"] = out.apply(honorarios, axis=1)
    out["Gasto limpieza"] = out["Ingreso limpieza"]
    out["Amenities"] = out.apply(amenities, axis=1)
    out["Total Gastos"] = out[["Comisión portal", "Honorarios Florit", "Gasto limpieza", "Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"] = out["Total ingresos"] - out["Comisión portal"]
    out = keep_original_dates_as_text(out, ["Fecha entrada","Fecha salida"])
    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas",
            "Ingreso alojamiento","Ingreso limpieza","Total ingresos",
            "Portal","Comisión portal","Honorarios Florit","Gasto limpieza",
            "Amenities","Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].sort_values(by=["Alojamiento","Fecha entrada"] if "Fecha entrada" in cols else ["Alojamiento"])

def process_case2(df):
    df = normalize_columns(df)
    out = df.copy()
    if {"Alojamiento","Portal","Comisión portal"}.issubset(out.columns):
        mask_apolo = out["Alojamiento"].isin({"APOLO 29","APOLO 197"})
        mask_book = out["Portal"].astype(str).str.lower().str.contains("booking", na=False)
        out.loc[mask_apolo & mask_book, "Comisión portal"] = out.loc[mask_apolo & mask_book, "Comisión portal"] * 1.21
    def honorarios(row):
        pct = case2_percent_amenities.get(row["Alojamiento"], (0.20, 0.0))[0]
        base = row["Ingreso alojamiento"] - row.get("IVA del alquiler", 0.0)
        return base * pct * 1.21
    def amenities(row):
        return case2_percent_amenities.get(row["Alojamiento"], (0.20, 0.0))[1]
    out["Honorarios Florit"] = out.apply(honorarios, axis=1)
    out["Gasto limpieza"] = out.get("Ingreso limpieza", 0.0)
    out["Amenities"] = out.apply(amenities, axis=1)
    out["Total Gastos"] = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"] = out["Total ingresos"] - out["Comisión portal"]
    out = keep_original_dates_as_text(out, ["Fecha entrada","Fecha salida"])
    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","IVA del alquiler",
            "Ingreso limpieza","Total ingresos","Portal","Comisión portal","Honorarios Florit","Gasto limpieza",
            "Amenities","Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].sort_values(by=["Alojamiento","Fecha entrada"] if "Fecha entrada" in cols else ["Alojamiento"])

def process_case3(df):
    df = normalize_columns(df)
    out = apply_booking_vat_on_commission(df, "Portal", "Comisión portal")
    def honorarios(row):
        pct = case3_cleaning_amenities.get(row["Alojamiento"], (0.20, None, None))[0]
        base = row["Ingreso alojamiento"] - row["Comisión portal"]
        return base * pct * 1.21
    def gasto_limpieza(row):
        return case3_cleaning_amenities.get(row["Alojamiento"], (0.20, 0.0, 0.0))[1]
    def amenities(row):
        return case3_cleaning_amenities.get(row["Alojamiento"], (0.20, 0.0, 0.0))[2]
    out = out.copy()
    out["Honorarios Florit"] = out.apply(honorarios, axis=1)
    out["Gasto limpieza"] = out.apply(gasto_limpieza, axis=1)
    out["Amenities"] = out.apply(amenities, axis=1)
    out["Total Gastos"] = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"] = out["Total ingresos"] - out["Comisión portal"]
    out = keep_original_dates_as_text(out, ["Fecha entrada","Fecha salida"])
    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","Ingreso limpieza","Total ingresos",
            "Portal","Comisión portal","Honorarios Florit","Gasto limpieza","Amenities","Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].sort_values(by=["Alojamiento","Fecha entrada"] if "Fecha entrada" in cols else ["Alojamiento"])

def process_case4(df):
    df = normalize_columns(df)
    out = df.copy()
    def honorarios(row):
        base = row["Ingreso alojamiento"] - row.get("IVA del alquiler", 0.0) - row["Comisión portal"]
        return base * 0.20
    out["Honorarios Florit"] = out.apply(honorarios, axis=1)
    out["Gasto limpieza"] = out.get("Ingreso limpieza", 0.0)
    out["Amenities"] = 0.0
    out["Total Gastos"] = out[["Comisión portal","Honorarios Florit"]].sum(axis=1) + out["Gasto limpieza"] + out["Amenities"]
    out["Pago al propietario"] = out["Ingreso alojamiento"] - out.get("IVA del alquiler", 0.0) - out["Comisión portal"] - out["Honorarios Florit"]
    out["Pago recibido"] = out["Ingreso alojamiento"] + out.get("Ingreso limpieza", 0.0) - out["Comisión portal"]
    out = keep_original_dates_as_text(out, ["Fecha entrada","Fecha salida"])
    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","IVA del alquiler",
            "Portal","Comisión portal","Honorarios Florit","Pago al propietario","Pago recibido","Ingreso limpieza"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].sort_values(by=["Alojamiento","Fecha entrada"] if "Fecha entrada" in cols else ["Alojamiento"])

def process_case5(df):
    df = normalize_columns(df)
    out = df.copy()
    def honorarios(row):
        pct = case5_percent_amenities.get(row["Alojamiento"], (0.20, 0.0))[0]
        base = row["Ingreso alojamiento"] - row.get("IVA del alquiler", 0.0) - row["Comisión portal"]
        return base * pct * 1.21
    def amenities(row):
        return case5_percent_amenities.get(row["Alojamiento"], (0.20, 0.0))[1]
    out["Honorarios Florit"] = out.apply(honorarios, axis=1)
    out["Gasto limpieza"] = out.get("Ingreso limpieza", 0.0)
    out["Amenities"] = out.apply(amenities, axis=1)
    out["Total Gastos"] = out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)
    out["Pago al propietario"] = out["Total ingresos"] - out["Total Gastos"]
    out["Pago recibido"] = out["Total ingresos"] - out["Comisión portal"]
    out = keep_original_dates_as_text(out, ["Fecha entrada","Fecha salida"])
    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","IVA del alquiler",
            "Ingreso limpieza","Total ingresos","Portal","Comisión portal","Honorarios Florit","Gasto limpieza",
            "Amenities","Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].sort_values(by=["Alojamiento","Fecha entrada"] if "Fecha entrada" in cols else ["Alojamiento"])

processors = {1: process_case1, 2: process_case2, 3: process_case3, 4: process_case4, 5: process_case5}

def detect_case(df):
    df_norm = normalize_columns(df)
    alos = set(df_norm.get("Alojamiento", pd.Series(dtype=str)).astype(str).str.upper().unique())
    scores = {
        1: len(alos & set(case1_props)),
        2: len(alos & set(case2_props)),
        3: len(alos & set(case3_props)),
        4: len(alos & set(case4_props)),
        5: len(alos & set(case5_props)),
    }
    best_case = max(scores, key=scores.get)
    return best_case, scores

# ----------------------
# Período y prorrateo
# ----------------------
def overlap_nights(checkin, checkout, start, end):
    if pd.isna(checkin) or pd.isna(checkout):
        return 0
    start = pd.to_datetime(start)
    end = pd.to_datetime(end) + pd.Timedelta(days=1)
    a = max(checkin, start)
    b = min(checkout, end)
    return max(0, (b - a).days)

def apply_period(result_df, start_date, end_date, prorate=True, limpieza_mode="prorratear", amenities_mode="prorratear"):
    df = result_df.copy()
    fe = pd.to_datetime(df.get("Fecha entrada"), errors="coerce", dayfirst=True)
    fs = pd.to_datetime(df.get("Fecha salida"), errors="coerce", dayfirst=True)
    total_noches = pd.to_numeric(df.get("Noches ocupadas"), errors="coerce").fillna(0)

    nights_in_period = [overlap_nights(ci, co, start_date, end_date) for ci, co in zip(fe, fs)]
    df["Noches en periodo"] = pd.Series(nights_in_period, index=df.index).astype(float)

    if not prorate:
        mask = df["Noches en periodo"] > 0
        out = df[mask].drop(columns=["Noches en periodo"])
        out["Fecha entrada"] = fe.dt.strftime("%d/%m/%Y")
        out["Fecha salida"] = fs.dt.strftime("%d/%m/%Y")
        return out

    ratio = np.where(total_noches.to_numpy() > 0, (df["Noches en periodo"] / total_noches).to_numpy(), 0.0)

    for col in ["Ingreso alojamiento","Total ingresos","Comisión portal","Honorarios Florit"]:
        if col in df.columns:
            df[col] = (pd.to_numeric(df[col], errors="coerce").fillna(0).to_numpy() * ratio).round(2)

    if "Ingreso limpieza" in df.columns and "Gasto limpieza" in df.columns:
        if limpieza_mode == "prorratear":
            df["Ingreso limpieza"] = (pd.to_numeric(df["Ingreso limpieza"], errors="coerce").fillna(0).to_numpy() * ratio).round(2)
            df["Gasto limpieza"] = df["Ingreso limpieza"]
        elif limpieza_mode == "salida":
            mask = (fs >= pd.to_datetime(start_date)) & (fs <= pd.to_datetime(end_date))
            df["Ingreso limpieza"] = np.where(mask, df["Ingreso limpieza"], 0.0)
            df["Gasto limpieza"] = df["Ingreso limpieza"]
        elif limpieza_mode == "entrada":
            mask = (fe >= pd.to_datetime(start_date)) & (fe <= pd.to_datetime(end_date))
            df["Ingreso limpieza"] = np.where(mask, df["Ingreso limpieza"], 0.0)
            df["Gasto limpieza"] = df["Ingreso limpieza"]

    if "Amenities" in df.columns:
        if amenities_mode == "prorratear":
            df["Amenities"] = (pd.to_numeric(df["Amenities"], errors="coerce").fillna(0).to_numpy() * ratio).round(2)
        elif amenities_mode == "salida":
            mask = (fs >= pd.to_datetime(start_date)) & (fs <= pd.to_datetime(end_date))
            df["Amenities"] = np.where(mask, df["Amenities"], 0.0)
        elif amenities_mode == "entrada":
            mask = (fe >= pd.to_datetime(start_date)) & (fe <= pd.to_datetime(end_date))
            df["Amenities"] = np.where(mask, df["Amenities"], 0.0)

    for c in ["Total Gastos","Pago al propietario","Pago recibido"]:
        if c in df.columns:
            df.drop(columns=[c], inplace=True, errors="ignore")

    if {"Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"}.issubset(df.columns):
        df["Total Gastos"] = df[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1).round(2)
    if {"Total ingresos","Total Gastos"}.issubset(df.columns):
        df["Pago al propietario"] = (df["Total ingresos"] - df["Total Gastos"]).round(2)
    if {"Total ingresos","Comisión portal"}.issubset(df.columns):
        df["Pago recibido"] = (df["Total ingresos"] - df["Comisión portal"]).round(2)

    df["Fecha entrada"] = fe.dt.strftime("%d/%m/%Y")
    df["Fecha salida"] = fs.dt.strftime("%d/%m/%Y")

    return df.drop(columns=["Noches en periodo"], errors="ignore")

# ----------------------
# UI
# ----------------------
st.title("📊 LIQUIDACIONES Automáticas (Casos 1–5)")
st.caption("Sube un Excel de Avantio (.xlsx). Detecta el caso y aplica reglas oficiales. Incluye filtro por período y prorrateo.")

with st.sidebar:
    st.header("Período a liquidar")
    col_a, col_b = st.columns(2)
    with col_a:
        start_date = st.date_input("Desde", value=date(date.today().year, date.today().month, 1))
    with col_b:
        end_date = st.date_input("Hasta", value=date(date.today().year, date.today().month, 28))

    st.markdown("**Reservas que cruzan meses**")
    prorate = st.checkbox("Prorratear importes por noches en el período (recomendado)", value=True)
    limpieza_mode = st.selectbox("Limpieza", ["prorratear","salida","entrada"], index=0)
    amenities_mode = st.selectbox("Amenities", ["prorratear","salida","entrada"], index=0)

file = st.file_uploader("Sube el archivo de reservas (.xlsx)", type=["xlsx"])

if file is not None:
    try:
        df_in = pd.read_excel(file, header=1)
    except Exception:
        df_in = pd.read_excel(file)

    st.write("Vista previa (primeras 12 filas):")
    st.dataframe(df_in.head(12), use_container_width=True)

    detected_case, scores = detect_case(df_in)
    st.info(f"**Caso detectado automáticamente: {detected_case}** | Recuento por caso: {scores}")
    case_choice = st.selectbox("Puedes elegir manualmente el caso", options=[1,2,3,4,5], index=[1,2,3,4,5].index(detected_case))

    if st.button("Generar liquidación"):
        processor = processors[case_choice]
        base_df = processor(df_in)
        final_df = apply_period(base_df, start_date, end_date, prorate=prorate, limpieza_mode=limpieza_mode, amenities_mode=amenities_mode)

        st.success(f"Liquidación generada (Caso {case_choice}) para el período {start_date.strftime('%d/%m/%Y')} – {end_date.strftime('%d/%m/%Y')}.")
        st.dataframe(final_df.head(50), use_container_width=True)

        excel_buffer = to_excel_grouped(final_df, group_col="Alojamiento", name=f"CASO {case_choice}")
        st.download_button(
            label="📥 Descargar Excel formateado",
            data=excel_buffer,
            file_name=f"Liquidacion_CASO_{case_choice}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.info("Sube un archivo para comenzar.")
